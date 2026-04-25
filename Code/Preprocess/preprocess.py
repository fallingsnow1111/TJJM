from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, cast

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PREPROCESS_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = PREPROCESS_DIR.parents[1]
DATASET_ROOT = PROJECT_ROOT / "Raw"
SUMMARY_JSON = PREPROCESS_DIR / "dataset_index_summary.json"
OUT_DIR = PREPROCESS_DIR / "output"


# Conversion factors to standard coal equivalent (10^4 tce per original unit in inventory row).
# Example: values in 10^4 ton multiply by tce/ton; values in 10^8 kWh multiply by 1.229.
ENERGY_TCE_FACTOR = {
	"Raw_Coal": 0.7143,
	"Cleaned_Coal": 0.9000,
	"Other_Washed_Coal": 0.2857,
	"Briquettes": 0.6000,
	"Coke": 0.9714,
	"Coke_Oven_Gas": 6.1430,
	"Other_Gas": 1.3300,
	"Other_Coking_Products": 1.1143,
	"Crude_Oil": 1.4286,
	"Gasoline": 1.4714,
	"Kerosene": 1.4714,
	"Diesel_Oil": 1.4571,
	"Fuel_Oil": 1.4286,
	"LPG": 1.7143,
	"Refinery_Gas": 1.5714,
	"Other_Petroleum_Products": 1.4286,
	"Natural_Gas": 13.3000,
	"Heat": 0.03412,
	"Electricity": 1.2290,
	"Other_Energy": 1.0000,
}

COAL_FUEL_HEADERS = {
	"Raw_Coal",
	"Cleaned_Coal",
	"Other_Washed_Coal",
	"Briquettes",
	"Coke",
	"Coke_Oven_Gas",
	"Other_Gas",
	"Other_Coking_Products",
}

PETROLEUM_HEADERS = {
	"Crude_Oil",
	"Gasoline",
	"Kerosene",
	"Diesel_Oil",
	"Fuel_Oil",
	"LPG",
	"Refinery_Gas",
	"Other_Petroleum_Products",
}

GAS_HEADERS = {"Natural_Gas"}

NON_FOSSIL_HEADERS = {"Heat", "Electricity", "Other_Energy"}

NATIONAL_ENERGY_TOTAL_PATH = DATASET_ROOT / "国家维度" / "能源相关" / "1978-2025能源消费总量.xlsx"


# Canonical province names aligned to inventory/MEIC naming.
PROVINCE_ALIAS = {
	"beijing": "Beijing",
	"tianjin": "Tianjin",
	"hebei": "Hebei",
	"shanxi": "Shanxi",
	"innermongolia": "InnerMongolia",
	"liaoning": "Liaoning",
	"jilin": "Jilin",
	"heilongjiang": "Heilongjiang",
	"shanghai": "Shanghai",
	"jiangsu": "Jiangsu",
	"zhejiang": "Zhejiang",
	"anhui": "Anhui",
	"fujian": "Fujian",
	"jiangxi": "Jiangxi",
	"shandong": "Shandong",
	"henan": "Henan",
	"hubei": "Hubei",
	"hunan": "Hunan",
	"guangdong": "Guangdong",
	"guangxi": "Guangxi",
	"hainan": "Hainan",
	"chongqing": "Chongqing",
	"sichuan": "Sichuan",
	"guizhou": "Guizhou",
	"yunnan": "Yunnan",
	"shaanxi": "Shaanxi",
	"gansu": "Gansu",
	"qinghai": "Qinghai",
	"ningxia": "Ningxia",
	"xinjiang": "Xinjiang",
	# Common variants
	"innermongoliaautonomousregion": "InnerMongolia",
	"innermongoliaregion": "InnerMongolia",
}

ZH_PROVINCE_ALIAS = {
	"北京": "Beijing",
	"北京市": "Beijing",
	"天津": "Tianjin",
	"天津市": "Tianjin",
	"河北": "Hebei",
	"河北省": "Hebei",
	"山西": "Shanxi",
	"山西省": "Shanxi",
	"内蒙古": "InnerMongolia",
	"内蒙古自治区": "InnerMongolia",
	"辽宁": "Liaoning",
	"辽宁省": "Liaoning",
	"吉林": "Jilin",
	"吉林省": "Jilin",
	"黑龙江": "Heilongjiang",
	"黑龙江省": "Heilongjiang",
	"上海": "Shanghai",
	"上海市": "Shanghai",
	"江苏": "Jiangsu",
	"江苏省": "Jiangsu",
	"浙江": "Zhejiang",
	"浙江省": "Zhejiang",
	"安徽": "Anhui",
	"安徽省": "Anhui",
	"福建": "Fujian",
	"福建省": "Fujian",
	"江西": "Jiangxi",
	"江西省": "Jiangxi",
	"山东": "Shandong",
	"山东省": "Shandong",
	"河南": "Henan",
	"河南省": "Henan",
	"湖北": "Hubei",
	"湖北省": "Hubei",
	"湖南": "Hunan",
	"湖南省": "Hunan",
	"广东": "Guangdong",
	"广东省": "Guangdong",
	"广西": "Guangxi",
	"广西壮族自治区": "Guangxi",
	"海南": "Hainan",
	"海南省": "Hainan",
	"重庆": "Chongqing",
	"重庆市": "Chongqing",
	"四川": "Sichuan",
	"四川省": "Sichuan",
	"贵州": "Guizhou",
	"贵州省": "Guizhou",
	"云南": "Yunnan",
	"云南省": "Yunnan",
	"陕西": "Shaanxi",
	"陕西省": "Shaanxi",
	"甘肃": "Gansu",
	"甘肃省": "Gansu",
	"青海": "Qinghai",
	"青海省": "Qinghai",
	"宁夏": "Ningxia",
	"宁夏回族自治区": "Ningxia",
	"新疆": "Xinjiang",
	"新疆维吾尔自治区": "Xinjiang",
}


def load_dataset_index() -> List[dict]:
	"""读取数据索引文件并返回条目列表。"""

	payload = json.loads(SUMMARY_JSON.read_text(encoding="utf-8"))
	items = payload["items"]
	for item in items:
		path = str(item.get("path", ""))
		resolved = ""
		if path and not Path(path).is_absolute():
			resolved = str(PROJECT_ROOT / path)
		elif path:
			resolved = path

		if not resolved:
			continue

		p = Path(resolved)
		if p.exists():
			item["path"] = str(p)
			continue

		# Backward-compatible remap for older index files using Dataset root.
		candidates = [
			resolved.replace("\\Dataset\\", "\\Raw\\"),
			resolved.replace("/Dataset/", "/Raw/"),
			resolved.replace("\\dataset\\", "\\Raw\\"),
			resolved.replace("/dataset/", "/Raw/"),
		]
		for cand in candidates:
			alt = Path(cand)
			if alt.exists():
				item["path"] = str(alt)
				break
	return items


def ensure_output_dir() -> None:
	"""确保输出目录存在。"""

	OUT_DIR.mkdir(parents=True, exist_ok=True)


def log_progress(message: str) -> None:
	"""打印带时间戳的进度信息，便于长任务观察运行状态。"""

	stamp = datetime.now().strftime("%H:%M:%S")
	print(f"[{stamp}] {message}", flush=True)


def normalize_province_name(name: str) -> Optional[str]:
	"""将中英文省份名称标准化为统一英文名称。"""

	if not isinstance(name, str):
		return None
	name = name.strip()
	if name in ZH_PROVINCE_ALIAS:
		return ZH_PROVINCE_ALIAS[name]
	compact = re.sub(r"[^A-Za-z]", "", name).lower()
	return PROVINCE_ALIAS.get(compact)


def extract_year_from_text(text: str) -> Optional[int]:
	"""从字符串中提取四位年份。"""

	if not isinstance(text, str):
		return None
	m = re.search(r"(19\d{2}|20\d{2})", text)
	return int(m.group(1)) if m else None


def parse_year_cell(v) -> Optional[int]:
	"""解析单元格中的年份值。"""

	if isinstance(v, (int, float)):
		if pd.isna(v):
			return None
		iv = int(v)
		if 1900 <= iv <= 2100:
			return iv
	elif isinstance(v, str):
		return extract_year_from_text(v)
	return None


def find_header_row(ws) -> int:
	"""在工作表中识别最可能的表头行。"""

	max_row = ws.max_row or 0
	max_col = ws.max_column or 0
	scan_rows = min(max_row, 15)

	best_row = 1
	best_year_hits = -1

	# Primary strategy: row with the largest count of year-like columns is the header row.
	for ridx in range(1, scan_rows + 1):
		hits = 0
		for c in range(2, max_col + 1):
			if parse_year_cell(ws.cell(row=ridx, column=c).value) is not None:
				hits += 1
		if hits > best_year_hits:
			best_year_hits = hits
			best_row = ridx

	if best_year_hits >= 5:
		return best_row

	# Fallback strategy for non-standard templates.
	for ridx in range(1, min(max_row, 25) + 1):
		val = ws.cell(row=ridx, column=1).value
		if isinstance(val, str) and val.strip() == "指标":
			return ridx

	return 1


def parse_year_headers(ws, header_row: int) -> Dict[int, int]:
	"""解析表头行，建立 年份->列号 的映射。"""

	year_to_col: Dict[int, int] = {}
	max_col = ws.max_column or 0
	for col in range(2, max_col + 1):
		v = ws.cell(row=header_row, column=col).value
		year = None
		year = parse_year_cell(v)
		if year is not None:
			year_to_col[year] = col
	return year_to_col


def to_float(v) -> Optional[float]:
	"""将常见数值格式安全转换为浮点数。"""

	if v is None:
		return None
	if isinstance(v, (int, float, np.number)):
		return float(v)
	if isinstance(v, str):
		s = v.strip().replace(",", "")
		if not s:
			return None
		try:
			return float(s)
		except ValueError:
			return None
	return None


def sort_panel_by_province_year(df: pd.DataFrame) -> pd.DataFrame:
	"""按省份和年份稳定排序，保证省内时间顺序。"""

	if "province" not in df.columns or "year" not in df.columns:
		return df
	return cast(
		pd.DataFrame,
		df.sort_values(["province", "year"], kind="mergesort").reset_index(drop=True),
	)


def ensure_metric_columns(df: pd.DataFrame, metric: str, include_imputed: bool = False) -> pd.DataFrame:
	"""确保变量及其来源标记列存在，减少重复样板代码。"""

	if metric not in df.columns:
		df[metric] = np.nan
	source_col = f"{metric}_source"
	if source_col not in df.columns:
		df[source_col] = np.nan
	if include_imputed:
		imputed_col = f"{metric}_is_imputed"
		if imputed_col not in df.columns:
			df[imputed_col] = 0
	return df


def read_meic_co2(items: List[dict]) -> pd.DataFrame:
	"""读取MEIC总排放并提取1990-2023省份CO2。"""

	meic_path = next((x["path"] for x in items if "MEIC" in x["path"]), None)
	if not meic_path:
		return pd.DataFrame(columns=["province", "year", "CO2", "CO2_source", "CO2_source_priority"])

	raw = pd.read_excel(meic_path, sheet_name="MEIC-China-CO2 total emissions", header=8)
	raw = raw.rename(columns={"Province": "province_raw", "Sector": "sector"})
	raw = raw[raw["sector"].astype(str).str.lower() == "total"].copy()

	year_cols = [c for c in raw.columns if isinstance(c, (int, float)) and 1900 <= int(c) <= 2100]
	if not year_cols:
		return pd.DataFrame(columns=["province", "year", "CO2", "CO2_source", "CO2_source_priority"])

	long_df = raw.melt(
		id_vars=["province_raw"],
		value_vars=year_cols,
		var_name="year",
		value_name="CO2",
	)
	long_df["year"] = long_df["year"].astype(int)
	long_df = long_df[(long_df["year"] >= 1990) & (long_df["year"] <= 2023)]
	long_df["province"] = long_df["province_raw"].map(normalize_province_name)
	long_df["CO2"] = pd.to_numeric(long_df["CO2"], errors="coerce")
	long_df = long_df.dropna(subset=["province", "CO2"])

	out = long_df[["province", "year", "CO2"]].copy()
	out["CO2_source"] = "MEIC_1990_2023"
	out["CO2_source_priority"] = 2
	return out


def build_co2_panel(items: List[dict]) -> pd.DataFrame:
	"""使用MEIC省级总排放构建CO2主面板。"""

	meic = read_meic_co2(items)

	combo = meic.sort_values(["province", "year", "CO2_source_priority"]).drop_duplicates(
		subset=["province", "year"],
		keep="first",
	)
	combo = combo[(combo["year"] >= 1990) & (combo["year"] <= 2023)].copy()

	return combo.sort_values(["province", "year"]).reset_index(drop=True)


def read_provincial_energy_inventory(items: List[dict]) -> pd.DataFrame:
	"""读取省级能源清单并按折标煤汇总Energy，同时计算煤/油/气/非化石占比。"""

	records: List[dict] = []

	energy_files = sorted(
		[
			x["path"]
			for x in items
			if Path(x["path"]).name.startswith("省级能源清单_") and Path(x["path"]).suffix.lower() == ".xlsx"
		]
	)

	total_files = len(energy_files)
	for fidx, path_str in enumerate(energy_files, start=1):
		log_progress(f"读取省级能源清单 {fidx}/{total_files}: {Path(path_str).name}")
		year = extract_year_from_text(Path(path_str).name)
		if year is None or year < 1990 or year > 2024:
			continue

		wb = load_workbook(path_str, read_only=False, data_only=True)
		try:
			for sheet_name in wb.sheetnames:
				if str(sheet_name).upper() == "NOTE":
					continue

				province = normalize_province_name(sheet_name)
				if not province:
					continue

				ws = wb[sheet_name]
				max_row = int(ws.max_row or 0)
				max_col = int(ws.max_column or 0)
				if max_row < 3 or max_col < 2:
					continue

				total_row = None
				for ridx in range(1, min(max_row, 120) + 1):
					v = ws.cell(row=ridx, column=1).value
					if isinstance(v, str) and v.strip().lower() == "total final consumption":
						total_row = ridx
						break

				if total_row is None:
					continue

				total_tce = 0.0
				coal_tce = 0.0
				oil_tce = 0.0
				gas_tce = 0.0
				nonfossil_tce = 0.0
				has_effective_value = False
				for c in range(2, max_col + 1):
					header = ws.cell(row=1, column=c).value
					if not isinstance(header, str):
						continue
					header = header.strip()
					factor = ENERGY_TCE_FACTOR.get(header)
					if factor is None:
						continue

					v = to_float(ws.cell(row=total_row, column=c).value)
					if v is None or not np.isfinite(v):
						continue

					contrib_tce = float(v) * float(factor)
					total_tce += contrib_tce
					has_effective_value = True
					if header in COAL_FUEL_HEADERS:
						coal_tce += contrib_tce
					elif header in PETROLEUM_HEADERS:
						oil_tce += contrib_tce
					elif header in GAS_HEADERS:
						gas_tce += contrib_tce
					elif header in NON_FOSSIL_HEADERS:
						nonfossil_tce += contrib_tce

				if not has_effective_value or total_tce <= 0:
					continue

				coal_share = (coal_tce / total_tce) * 100.0 if total_tce > 0 else np.nan
				oil_share = (oil_tce / total_tce) * 100.0 if total_tce > 0 else np.nan
				gas_share = (gas_tce / total_tce) * 100.0 if total_tce > 0 else np.nan
				nonfossil_share = (nonfossil_tce / total_tce) * 100.0 if total_tce > 0 else np.nan

				records.append(
					{
						"province": province,
						"year": int(year),
						"Energy": float(total_tce),
						"CoalTCE": float(coal_tce),
						"OilTCE": float(oil_tce),
						"GasTCE": float(gas_tce),
						"NonFossilTCE": float(nonfossil_tce),
						"Energy_source": "Provincial_energy_inventory_1997_2022_total_final_consumption_tce_converted",
						"Energy_is_national_proxy": 0,
						"CoalShare": float(coal_share),
						"OilShare": float(oil_share),
						"GasShare": float(gas_share),
						"NonFossilShare": float(nonfossil_share),
						"CoalShare_source": "Provincial_coal_share_from_total_final_consumption_tce_converted",
						"CoalShare_is_national_proxy": 0,
						"OilShare_source": "Provincial_oil_share_from_total_final_consumption_tce_converted",
						"OilShare_is_national_proxy": 0,
						"GasShare_source": "Provincial_gas_share_from_total_final_consumption_tce_converted",
						"GasShare_is_national_proxy": 0,
						"NonFossilShare_source": "Provincial_nonfossil_share_from_total_final_consumption_tce_converted",
						"NonFossilShare_is_national_proxy": 0,
					}
				)
		finally:
			wb.close()

	df = pd.DataFrame(records)
	if df.empty:
		return pd.DataFrame(
			columns=[
				"province",
				"year",
				"Energy",
				"CoalTCE",
				"OilTCE",
				"GasTCE",
				"NonFossilTCE",
				"Energy_source",
				"Energy_is_national_proxy",
				"CoalShare",
				"OilShare",
				"GasShare",
				"NonFossilShare",
				"CoalShare_source",
				"CoalShare_is_national_proxy",
				"OilShare_source",
				"OilShare_is_national_proxy",
				"GasShare_source",
				"GasShare_is_national_proxy",
				"NonFossilShare_source",
				"NonFossilShare_is_national_proxy",
			]
		)

	df = df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last")
	return df.reset_index(drop=True)


def _fit_transform_series(series: pd.Series, mode: str) -> pd.Series:
	if mode == "log":
		return cast(pd.Series, np.log(series.where(series > 0)))
	if mode == "logit":
		clipped = series.clip(lower=0.01, upper=99.99)
		return cast(pd.Series, np.log(clipped / (100.0 - clipped)))
	return series.copy()


def _inverse_transform_series(series: pd.Series, mode: str) -> pd.Series:
	if mode == "log":
		return cast(pd.Series, np.exp(series))
	if mode == "logit":
		return cast(pd.Series, 100.0 / (1.0 + np.exp(-series)))
	return series.copy()


def fill_series_provincial_only(
	panel: pd.DataFrame,
	metric: str,
	transform_mode: str,
	interpolation_source: str,
	polyfit_source: str,
	min_valid: float,
	max_valid: Optional[float] = None,
	observed_col: Optional[str] = None,
	observed_source: Optional[str] = None,
) -> pd.DataFrame:
	"""省内补全：中间缺口插值，前后缺口多项式拟合。"""

	df = panel.copy()
	df = ensure_metric_columns(df, metric, include_imputed=True)

	source_col = f"{metric}_source"
	imputed_col = f"{metric}_is_imputed"
	df[source_col] = df[source_col].astype(object)

	if observed_col and observed_col in df.columns:
		obs_mask = df[observed_col].notna()
		df.loc[obs_mask, metric] = pd.to_numeric(df.loc[obs_mask, observed_col], errors="coerce")
		if observed_source:
			df.loc[obs_mask, source_col] = observed_source
		df.loc[obs_mask, imputed_col] = 0
		df = df.drop(columns=[observed_col], errors="ignore")

	for _, idx in df.groupby("province").groups.items():
		sub = cast(pd.DataFrame, df.loc[list(idx), :].sort_values("year").copy())
		sub[metric] = pd.to_numeric(sub[metric], errors="coerce")
		sub.loc[sub[metric] <= min_valid, metric] = np.nan
		if max_valid is not None:
			sub.loc[sub[metric] >= max_valid, metric] = np.nan

		years = pd.to_numeric(sub["year"], errors="coerce")
		if years.isna().all():
			continue

		base = sub[metric].copy()
		trans = _fit_transform_series(base, transform_mode)
		trans.index = years.to_numpy()

		# 1) 中间缺口插值
		interp_year = trans.interpolate(method="index", limit_area="inside")
		interp = pd.Series(interp_year.to_numpy(), index=sub.index)
		inside_mask = base.isna() & interp.notna()
		if inside_mask.any():
			filled_inside = _inverse_transform_series(interp, transform_mode)
			sub.loc[inside_mask, metric] = filled_inside.loc[inside_mask]
			sub.loc[inside_mask, source_col] = interpolation_source
			sub.loc[inside_mask, imputed_col] = 1

		# 2) 前后缺口多项式拟合
		updated = pd.to_numeric(sub[metric], errors="coerce")
		obs_mask = updated.notna()
		if int(obs_mask.sum()) > 0:
			x_obs = years.loc[obs_mask].to_numpy(dtype=float)
			y_obs = _fit_transform_series(updated.loc[obs_mask], transform_mode).to_numpy(dtype=float)
			deg = int(min(2, len(x_obs) - 1))
			if deg >= 0:
				coef = np.polyfit(x_obs, y_obs, deg)
				poly = np.poly1d(coef)
				first_obs_year = float(np.min(x_obs))
				last_obs_year = float(np.max(x_obs))

				missing_mask = updated.isna()
				edge_mask = missing_mask & ((years < first_obs_year) | (years > last_obs_year))
				if edge_mask.any():
					x_edge = years.loc[edge_mask].to_numpy(dtype=float)
					y_edge = pd.Series(poly(x_edge), index=sub.index[edge_mask])
					filled_edge = _inverse_transform_series(y_edge, transform_mode)
					sub.loc[edge_mask, metric] = filled_edge
					sub.loc[edge_mask, source_col] = polyfit_source
					sub.loc[edge_mask, imputed_col] = 1

		if max_valid is not None:
			sub[metric] = sub[metric].clip(lower=min_valid, upper=max_valid)
		else:
			sub[metric] = sub[metric].clip(lower=min_valid)

		df.loc[sub.index, [metric, source_col, imputed_col]] = sub[[metric, source_col, imputed_col]]

	df[metric] = pd.to_numeric(df[metric], errors="coerce")
	df[imputed_col] = pd.to_numeric(df[imputed_col], errors="coerce").fillna(0)
	return sort_panel_by_province_year(df)


def read_provincial_macro_series(
	items: List[dict],
	path_keywords: Iterable[str],
	value_name: str,
	source_name: str,
	year_min: int = 1990,
	year_max: int = 2023,
) -> pd.DataFrame:
	"""读取省级宏观变量（第一列指标、第二列省份、后续列年份）的通用格式。"""

	path_str = select_single_sheet_path_by_keywords(items, path_keywords)
	if not path_str:
		return pd.DataFrame(columns=["province", "year", value_name, f"{value_name}_source", f"{value_name}_is_national_proxy"])

	wb = load_workbook(path_str, read_only=False, data_only=True)
	try:
		ws = wb[wb.sheetnames[0]]
		year_cols: Dict[int, int] = {}
		for col in range(3, int(ws.max_column or 0) + 1):
			year = parse_year_cell(ws.cell(row=1, column=col).value)
			if year is not None and year_min <= int(year) <= year_max:
				year_cols[int(year)] = col

		if not year_cols:
			return pd.DataFrame(columns=["province", "year", value_name, f"{value_name}_source", f"{value_name}_is_national_proxy"])

		records: List[dict] = []
		for ridx in range(2, int(ws.max_row or 0) + 1):
			province_raw = ws.cell(row=ridx, column=2).value
			if not isinstance(province_raw, str):
				continue
			province = normalize_province_name(province_raw)
			if not province:
				continue

			for year, col in sorted(year_cols.items()):
				v = to_float(ws.cell(row=ridx, column=col).value)
				if v is None:
					continue
				records.append(
					{
						"province": province,
						"year": int(year),
						value_name: float(v),
						f"{value_name}_source": source_name,
						f"{value_name}_is_national_proxy": 0,
					}
				)
	finally:
		wb.close()

	df = pd.DataFrame(records)
	if df.empty:
		return pd.DataFrame(columns=["province", "year", value_name, f"{value_name}_source", f"{value_name}_is_national_proxy"])

	df = df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last")
	return df.reset_index(drop=True)


def read_provincial_gdp(items: List[dict]) -> pd.DataFrame:
	"""读取省级GDP观测值。"""

	return read_provincial_macro_series(
		items,
		path_keywords=["省份维度", "经济相关", "1949-2024省级GDP"],
		value_name="GDP",
		source_name="Provincial_GDP_1949_2024_observed",
		year_min=1990,
		year_max=2023,
	)


def read_provincial_population(items: List[dict]) -> pd.DataFrame:
	"""读取省级人口观测值。"""

	return read_provincial_macro_series(
		items,
		path_keywords=["省份维度", "人口相关", "1949-2024省级人口"],
		value_name="Population",
		source_name="Provincial_population_1949_2024_observed",
		year_min=1990,
		year_max=2023,
	)


def log_observed_coverage(name: str, df: pd.DataFrame, value_col: str) -> None:
	"""输出变量覆盖范围，作为是否需要补全的诊断依据。"""

	if df.empty or value_col not in df.columns:
		log_progress(f"{name} 省级观测为空")
		return
	obs = df[["province", "year", value_col]].copy()
	obs[value_col] = pd.to_numeric(obs[value_col], errors="coerce")
	obs = obs.dropna(subset=[value_col])
	if obs.empty:
		log_progress(f"{name} 省级观测为空")
		return
	log_progress(
		f"{name} 省级覆盖: 省份{obs['province'].nunique()}个, 年份{int(obs['year'].min())}-{int(obs['year'].max())}, 观测{len(obs)}条"
	)


def fill_positive_with_provincial_only(
	panel: pd.DataFrame,
	value_col: str,
	interpolation_source: str,
	polyfit_source: str,
) -> pd.DataFrame:
	"""正值变量补全：中间缺口插值，前后缺口多项式拟合。"""

	return fill_series_provincial_only(
		panel=panel,
		metric=value_col,
		transform_mode="log",
		interpolation_source=interpolation_source,
		polyfit_source=polyfit_source,
		min_valid=0.0,
		max_valid=None,
	)


def select_single_sheet_path_by_keywords(items: List[dict], keywords: Iterable[str]) -> Optional[str]:
	"""按多个关键词同时匹配选择单表文件路径。"""

	for it in items:
		if it.get("sheet_count") != 1:
			continue
		path = str(it.get("path", ""))
		if all(k in path for k in keywords):
			return path
	return None


def read_provincial_industry_share(items: List[dict]) -> pd.DataFrame:
	"""读取省级第二产业占比观测值。"""

	path_str = select_single_sheet_path_by_keywords(items, ["省份维度", "经济相关", "省份第二产业占比"])
	if not path_str:
		return pd.DataFrame(columns=["province", "year", "Industry", "Industry_source", "Industry_is_national_proxy"])

	wb = load_workbook(path_str, read_only=False, data_only=True)
	try:
		ws = wb[wb.sheetnames[0]]

		year_cols: Dict[int, int] = {}
		for col in range(3, int(ws.max_column or 0) + 1):
			year = parse_year_cell(ws.cell(row=1, column=col).value)
			if year is not None and 1990 <= int(year) <= 2022:
				year_cols[int(year)] = col

		if not year_cols:
			return pd.DataFrame(columns=["province", "year", "Industry", "Industry_source", "Industry_is_national_proxy"])

		records: List[dict] = []
		last_indicator: Optional[str] = None
		for ridx in range(2, int(ws.max_row or 0) + 1):
			indicator_cell = ws.cell(row=ridx, column=1).value
			if isinstance(indicator_cell, str) and indicator_cell.strip():
				last_indicator = indicator_cell.strip()
			indicator = last_indicator
			province_raw = ws.cell(row=ridx, column=2).value
			if not isinstance(indicator, str):
				continue
			if not isinstance(province_raw, str):
				continue
			if "第二产业" not in indicator or "比重" not in indicator:
				continue

			province = normalize_province_name(province_raw)
			if not province:
				continue

			for year, col in sorted(year_cols.items()):
				v = to_float(ws.cell(row=ridx, column=col).value)
				if v is None:
					continue
				records.append(
					{
						"province": province,
						"year": int(year),
						"Industry": float(v),
						"Industry_source": "Provincial_industry_share_1996_2024_observed",
						"Industry_is_national_proxy": 0,
					}
				)
	finally:
		wb.close()

	df = pd.DataFrame(records)
	if df.empty:
		return pd.DataFrame(columns=["province", "year", "Industry", "Industry_source", "Industry_is_national_proxy"])

	df = df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last")
	return df.reset_index(drop=True)


def fill_share_with_provincial_first(
	panel: pd.DataFrame,
	metric: str,
	interpolation_source: str,
	polyfit_source: str,
	observed_col: Optional[str] = None,
	observed_source: Optional[str] = None,
) -> pd.DataFrame:
	"""占比类变量补全：中间缺口插值，前后缺口多项式拟合。"""

	return fill_series_provincial_only(
		panel=panel,
		metric=metric,
		transform_mode="logit",
		interpolation_source=interpolation_source,
		polyfit_source=polyfit_source,
		min_valid=0.0,
		max_valid=100.0,
		observed_col=observed_col,
		observed_source=observed_source,
	)


def fill_industry_with_interpolation_fit_and_anchor(
	panel: pd.DataFrame,
) -> pd.DataFrame:
	"""按省级规则补全Industry。"""

	return fill_share_with_provincial_first(
		panel=panel,
		metric="Industry",
		interpolation_source="Provincial_industry_logit_interpolation",
		polyfit_source="Provincial_industry_logit_polynomial_fit",
	)


def read_provincial_urbanization_share(items: List[dict]) -> pd.DataFrame:
	"""读取省级城镇化率（城镇人口占比）观测值。"""

	path_str = select_single_sheet_path_by_keywords(items, ["省份维度", "人口相关", "城镇人口所占比重"])
	if not path_str:
		return pd.DataFrame(columns=["province", "year", "Urbanization_prov"])

	wb = load_workbook(path_str, read_only=False, data_only=True)
	try:
		ws = wb[wb.sheetnames[0]]

		year_cols: Dict[int, int] = {}
		for col in range(3, int(ws.max_column or 0) + 1):
			year = parse_year_cell(ws.cell(row=1, column=col).value)
			if year is not None and 1990 <= int(year) <= 2022:
				year_cols[int(year)] = col

		if not year_cols:
			return pd.DataFrame(columns=["province", "year", "Urbanization_prov"])

		records: List[dict] = []
		last_indicator: Optional[str] = None
		for ridx in range(2, int(ws.max_row or 0) + 1):
			indicator_cell = ws.cell(row=ridx, column=1).value
			if isinstance(indicator_cell, str) and indicator_cell.strip():
				last_indicator = indicator_cell.strip()
			indicator = last_indicator
			province_raw = ws.cell(row=ridx, column=2).value
			if not isinstance(indicator, str):
				continue
			if not isinstance(province_raw, str):
				continue
			if "城镇人口" not in indicator or "比重" not in indicator:
				continue

			province = normalize_province_name(province_raw)
			if not province:
				continue

			for year, col in sorted(year_cols.items()):
				v = to_float(ws.cell(row=ridx, column=col).value)
				if v is None:
					continue
				records.append(
					{
						"province": province,
						"year": int(year),
						"Urbanization_prov": float(v),
					}
				)
	finally:
		wb.close()

	df = pd.DataFrame(records)
	if df.empty:
		return pd.DataFrame(columns=["province", "year", "Urbanization_prov"])

	df = df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last")
	return df.reset_index(drop=True)


def fill_urbanization_with_interpolation_fit_and_anchor(
	panel: pd.DataFrame,
) -> pd.DataFrame:
	"""按省级规则补全Urbanization。"""

	return fill_share_with_provincial_first(
		panel=panel,
		metric="Urbanization",
		interpolation_source="Provincial_urbanization_logit_interpolation",
		polyfit_source="Provincial_urbanization_logit_polynomial_fit",
		observed_col="Urbanization_prov",
		observed_source="Provincial_urbanization_share_1990_2025_observed",
	)


def fill_energy_with_interpolation_and_fit(
	panel: pd.DataFrame,
) -> pd.DataFrame:
	"""按省级规则补全Energy。"""

	return fill_positive_with_provincial_only(
		panel=panel,
		value_col="Energy",
		interpolation_source="Provincial_energy_log_interpolation",
		polyfit_source="Provincial_energy_log_polynomial_fit",
	)


def fill_coal_share_with_interpolation_and_fit(
	panel: pd.DataFrame,
) -> pd.DataFrame:
	"""按省级规则补全CoalShare。"""

	return fill_share_with_provincial_first(
		panel=panel,
		metric="CoalShare",
		interpolation_source="Provincial_coal_share_logit_interpolation",
		polyfit_source="Provincial_coal_share_logit_polynomial_fit",
	)


def fill_oil_share_with_interpolation_and_fit(
	panel: pd.DataFrame,
) -> pd.DataFrame:
	"""按省级规则补全OilShare。"""

	return fill_share_with_provincial_first(
		panel=panel,
		metric="OilShare",
		interpolation_source="Provincial_oil_share_logit_interpolation",
		polyfit_source="Provincial_oil_share_logit_polynomial_fit",
	)


def fill_gas_share_with_interpolation_and_fit(
	panel: pd.DataFrame,
) -> pd.DataFrame:
	"""按省级规则补全GasShare。"""

	return fill_share_with_provincial_first(
		panel=panel,
		metric="GasShare",
		interpolation_source="Provincial_gas_share_logit_interpolation",
		polyfit_source="Provincial_gas_share_logit_polynomial_fit",
	)


def fill_nonfossil_share_with_interpolation_and_fit(
	panel: pd.DataFrame,
) -> pd.DataFrame:
	"""按省级规则补全NonFossilShare。"""

	return fill_share_with_provincial_first(
		panel=panel,
		metric="NonFossilShare",
		interpolation_source="Provincial_nonfossil_share_logit_interpolation",
		polyfit_source="Provincial_nonfossil_share_logit_polynomial_fit",
	)


def read_national_energy_total(path: Path) -> pd.DataFrame:
	"""读取全国能源消费总量文件并返回 year-total_energy 两列。"""

	if not path.exists():
		return pd.DataFrame(columns=["year", "national_energy_total"])

	try:
		df = pd.read_excel(path, header=2)
	except Exception:
		return pd.DataFrame(columns=["year", "national_energy_total"])

	if df.empty:
		return pd.DataFrame(columns=["year", "national_energy_total"])

	indicator_col = str(df.columns[0])
	row_mask = df[indicator_col].astype(str).str.contains("能源消费总量", na=False)
	if not row_mask.any():
		return pd.DataFrame(columns=["year", "national_energy_total"])

	row = df.loc[row_mask].iloc[0]
	records: List[dict] = []
	for col in df.columns[1:]:
		year = parse_year_cell(str(col)) if isinstance(col, str) else parse_year_cell(col)
		if year is None:
			continue
		v = to_float(row[col])
		if v is None or not np.isfinite(v):
			continue
		records.append({"year": int(year), "national_energy_total": float(v)})

	out = pd.DataFrame(records)
	if out.empty:
		return pd.DataFrame(columns=["year", "national_energy_total"])

	out = out[(out["year"] >= 1990) & (out["year"] <= 2025)]
	out = out.sort_values("year", kind="stable").drop_duplicates(subset=["year"], keep="last")
	return out.reset_index(drop=True)


def build_national_energy_validation(panel_core: pd.DataFrame) -> tuple[pd.DataFrame, Dict[str, float]]:
	"""对比省级Energy加总与全国能源消费总量，生成校验表与统计。"""

	province_sum = (
		panel_core.groupby("year", as_index=False)["Energy"]
		.sum()
		.rename(columns={"Energy": "province_energy_sum"})
	)
	national_df = read_national_energy_total(NATIONAL_ENERGY_TOTAL_PATH)

	if national_df.empty:
		out = province_sum.copy()
		out["national_energy_total"] = np.nan
		out["ratio_actual_over_province_sum"] = np.nan
		out["diff_actual_minus_province_sum"] = np.nan
		return out, {}

	out = province_sum.merge(national_df, on="year", how="left")
	out["ratio_actual_over_province_sum"] = out["national_energy_total"] / out["province_energy_sum"]
	out["diff_actual_minus_province_sum"] = out["national_energy_total"] - out["province_energy_sum"]

	overlap = out.dropna(subset=["national_energy_total"]).copy()
	if overlap.empty:
		return out, {}

	den = np.maximum(overlap["national_energy_total"].abs().to_numpy(dtype=float), 1e-8)
	mape = np.mean(np.abs((overlap["province_energy_sum"].to_numpy(dtype=float) - overlap["national_energy_total"].to_numpy(dtype=float)) / den))
	stats = {
		"overlap_year_count": int(len(overlap)),
		"overlap_year_min": int(overlap["year"].min()),
		"overlap_year_max": int(overlap["year"].max()),
		"corr": float(overlap["province_energy_sum"].corr(overlap["national_energy_total"])),
		"mape": float(mape),
		"mean_ratio_actual_over_province_sum": float(overlap["ratio_actual_over_province_sum"].mean()),
	}
	return out, stats


def read_provincial_transport_mileage_and_private_cars(items: List[dict]) -> pd.DataFrame:
	"""读取省级交通数据并提取HighwayMileage与PrivateCars观测值。"""

	path_str = select_single_sheet_path_by_keywords(items, ["省份维度", "交通相关", "里程和汽车拥有量省份数据"])
	transport_path: Optional[Path] = None
	if path_str:
		candidate = PROJECT_ROOT / path_str
		if candidate.exists():
			transport_path = candidate

	if transport_path is None:
		fallback_paths = sorted((DATASET_ROOT / "省份维度" / "交通相关").glob("*里程和汽车拥有量省份数据*.xlsx"))
		if fallback_paths:
			transport_path = fallback_paths[0]

	if transport_path is None:
		return pd.DataFrame(
			columns=[
				"province",
				"year",
				"HighwayMileage",
				"HighwayMileage_source",
				"HighwayMileage_is_national_proxy",
				"PrivateCars",
				"PrivateCars_source",
				"PrivateCars_is_national_proxy",
			]
		)

	wb = load_workbook(transport_path, read_only=False, data_only=True)
	try:
		ws = wb[wb.sheetnames[0]]

		year_cols: Dict[int, int] = {}
		for col in range(3, int(ws.max_column or 0) + 1):
			year = parse_year_cell(ws.cell(row=1, column=col).value)
			if year is not None and 1990 <= int(year) <= 2023:
				year_cols[int(year)] = col

		if not year_cols:
			return pd.DataFrame(
				columns=[
					"province",
					"year",
					"HighwayMileage",
					"HighwayMileage_source",
					"HighwayMileage_is_national_proxy",
					"PrivateCars",
					"PrivateCars_source",
					"PrivateCars_is_national_proxy",
				]
			)

		highway_records: List[dict] = []
		car_records: List[dict] = []
		last_indicator: Optional[str] = None
		used_railway_mileage = False
		used_civil_car = False

		for ridx in range(2, int(ws.max_row or 0) + 1):
			indicator_cell = ws.cell(row=ridx, column=1).value
			if isinstance(indicator_cell, str) and indicator_cell.strip():
				last_indicator = indicator_cell.strip()
			indicator = last_indicator
			province_raw = ws.cell(row=ridx, column=2).value
			if not isinstance(indicator, str):
				continue
			if not isinstance(province_raw, str):
				continue

			indicator_compact = re.sub(r"\s+", "", indicator)
			province = normalize_province_name(province_raw)
			if not province:
				continue

			is_highway = "公路" in indicator_compact and "里程" in indicator_compact
			is_railway_fallback = "铁路" in indicator_compact and "里程" in indicator_compact
			is_private_cars = "私家车" in indicator_compact and "拥有量" in indicator_compact
			is_civil_car_fallback = "民用汽车" in indicator_compact and "拥有量" in indicator_compact

			for year, col in sorted(year_cols.items()):
				v = to_float(ws.cell(row=ridx, column=col).value)
				if v is None:
					continue

				if is_highway or is_railway_fallback:
					source_name = "Provincial_highway_mileage_1990_2024_observed"
					if is_railway_fallback and not is_highway:
						source_name = "Provincial_railway_mileage_1990_2024_used_as_highway_proxy"
						used_railway_mileage = True
					highway_records.append(
						{
							"province": province,
							"year": int(year),
							"HighwayMileage": float(v),
							"HighwayMileage_source": source_name,
							"HighwayMileage_is_national_proxy": 0,
						}
					)

				if is_private_cars or is_civil_car_fallback:
					source_name = "Provincial_private_cars_1990_2024_observed"
					if is_civil_car_fallback and not is_private_cars:
						source_name = "Provincial_civil_car_1990_2024_used_as_private_cars_proxy"
						used_civil_car = True
					car_records.append(
						{
							"province": province,
							"year": int(year),
							"PrivateCars": float(v),
							"PrivateCars_source": source_name,
							"PrivateCars_is_national_proxy": 0,
						}
					)
	finally:
		wb.close()

	if used_railway_mileage:
		log_progress("交通口径提示：未发现公路里程，当前使用铁路营业里程作为HighwayMileage代理。")
	if used_civil_car:
		log_progress("交通口径提示：未发现私家车数量，当前使用民用汽车拥有量作为PrivateCars代理。")

	highway_df = pd.DataFrame(highway_records)
	car_df = pd.DataFrame(car_records)

	if highway_df.empty:
		highway_df = pd.DataFrame(
			columns=["province", "year", "HighwayMileage", "HighwayMileage_source", "HighwayMileage_is_national_proxy"]
		)
	else:
		highway_df = highway_df.sort_values(["province", "year"]).drop_duplicates(
			subset=["province", "year"],
			keep="last",
		)

	if car_df.empty:
		car_df = pd.DataFrame(columns=["province", "year", "PrivateCars", "PrivateCars_source", "PrivateCars_is_national_proxy"])
	else:
		car_df = car_df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last")

	out = highway_df.merge(car_df, on=["province", "year"], how="outer")
	if out.empty:
		return pd.DataFrame(
			columns=[
				"province",
				"year",
				"HighwayMileage",
				"HighwayMileage_source",
				"HighwayMileage_is_national_proxy",
				"PrivateCars",
				"PrivateCars_source",
				"PrivateCars_is_national_proxy",
			]
		)

	return out.sort_values(["province", "year"]).reset_index(drop=True)


def log_mean(a: float, b: float) -> Optional[float]:
	"""计算对数均值，用于 LMDI 加法分解。"""

	if a <= 0 or b <= 0:
		return None
	if abs(a - b) < 1e-12:
		return float(a)
	la = np.log(a)
	lb = np.log(b)
	if abs(la - lb) < 1e-12:
		return float(a)
	return float((a - b) / (la - lb))


def compute_lmdi_time(panel: pd.DataFrame) -> pd.DataFrame:
	"""按省份逐年计算四因子 LMDI: CO2 = Population * A * B * C。"""

	rows: List[dict] = []
	panel_sorted = sort_panel_by_province_year(panel)

	for province, g in panel_sorted.groupby("province"):
		g = cast(pd.DataFrame, g.reset_index(drop=True))
		for i in range(1, len(g)):
			prev = g.iloc[i - 1]
			curr = g.iloc[i]

			e0 = float(prev["CO2"])
			et = float(curr["CO2"])
			p0 = float(prev["Population"])
			pt = float(curr["Population"])
			a0 = float(prev["A"])
			at = float(curr["A"])
			b0 = float(prev["B"])
			bt = float(curr["B"])
			c0 = float(prev["C"])
			ct = float(curr["C"])

			vals = [e0, et, p0, pt, a0, at, b0, bt, c0, ct]
			if not all(np.isfinite(v) and v > 0 for v in vals):
				continue

			lm = log_mean(et, e0)
			if lm is None:
				continue

			d_p = lm * np.log(pt / p0)
			d_a = lm * np.log(at / a0)
			d_b = lm * np.log(bt / b0)
			d_c = lm * np.log(ct / c0)
			d_co2 = et - e0
			resid = d_co2 - (d_p + d_a + d_b + d_c)
			resid_ratio_abs = abs(resid) / max(abs(d_co2), 1e-12)

			rows.append(
				{
					"province": province,
					"year": int(curr["year"]),
					"base_year": int(prev["year"]),
					"delta_CO2": d_co2,
					"delta_P": d_p,
					"delta_A": d_a,
					"delta_B": d_b,
					"delta_C": d_c,
					"lmdi_residual": resid,
					"resid_ratio": resid_ratio_abs,
					"lmdi_residual_abs_ratio": resid_ratio_abs,
				}
			)

	return pd.DataFrame(rows)


def main() -> None:
	"""执行数据预处理主流程并写出结果文件。"""

	log_progress("开始执行预处理流程")
	ensure_output_dir()
	items = load_dataset_index()
	log_progress(f"已加载数据索引，共 {len(items)} 个条目")

	log_progress("阶段1/7：构建CO2主面板")
	co2_panel = build_co2_panel(items)
	log_progress(f"CO2主面板完成，记录数 {len(co2_panel)}")

	log_progress("阶段2/7：读取省级GDP与Population")
	prov_gdp = read_provincial_gdp(items)
	prov_population = read_provincial_population(items)
	log_observed_coverage("GDP", prov_gdp, "GDP")
	log_observed_coverage("Population", prov_population, "Population")
	panel = co2_panel.copy()
	if not prov_gdp.empty:
		panel = panel.merge(prov_gdp, on=["province", "year"], how="left")
	if not prov_population.empty:
		panel = panel.merge(prov_population, on=["province", "year"], how="left")

	log_progress("阶段3/7：补全GDP与Population（省内插值+前后多项式拟合）")
	panel = fill_positive_with_provincial_only(
		panel,
		value_col="GDP",
		interpolation_source="Provincial_GDP_log_interpolation",
		polyfit_source="Provincial_GDP_log_polynomial_fit",
	)
	panel = fill_positive_with_provincial_only(
		panel,
		value_col="Population",
		interpolation_source="Provincial_population_log_interpolation",
		polyfit_source="Provincial_population_log_polynomial_fit",
	)

	log_progress("阶段4/7：补全Industry")
	prov_industry = read_provincial_industry_share(items)
	log_observed_coverage("Industry", prov_industry, "Industry")
	if not prov_industry.empty:
		panel = panel.merge(prov_industry, on=["province", "year"], how="left")
	panel = ensure_metric_columns(panel, "Industry", include_imputed=True)

	panel = fill_industry_with_interpolation_fit_and_anchor(panel)

	log_progress("阶段5/7：补全Urbanization")
	prov_urbanization = read_provincial_urbanization_share(items)
	log_observed_coverage("Urbanization", prov_urbanization, "Urbanization_prov")
	if not prov_urbanization.empty:
		panel = panel.merge(prov_urbanization, on=["province", "year"], how="left")
	panel = ensure_metric_columns(panel, "Urbanization", include_imputed=True)

	panel = fill_urbanization_with_interpolation_fit_and_anchor(panel)

	log_progress("阶段6/7：补全Energy与能源结构占比（Coal/Oil/Gas/NonFossil）")
	prov_energy = read_provincial_energy_inventory(items)
	log_observed_coverage("Energy", prov_energy, "Energy")
	log_observed_coverage("CoalShare", prov_energy, "CoalShare")
	log_observed_coverage("OilShare", prov_energy, "OilShare")
	log_observed_coverage("GasShare", prov_energy, "GasShare")
	log_observed_coverage("NonFossilShare", prov_energy, "NonFossilShare")
	if not prov_energy.empty:
		panel = panel.merge(prov_energy, on=["province", "year"], how="left")
	panel = ensure_metric_columns(panel, "Energy", include_imputed=True)
	panel = ensure_metric_columns(panel, "CoalShare", include_imputed=True)
	panel = ensure_metric_columns(panel, "OilShare", include_imputed=True)
	panel = ensure_metric_columns(panel, "GasShare", include_imputed=True)
	panel = ensure_metric_columns(panel, "NonFossilShare", include_imputed=True)

	panel = fill_energy_with_interpolation_and_fit(panel)
	panel = fill_coal_share_with_interpolation_and_fit(panel)
	panel = fill_oil_share_with_interpolation_and_fit(panel)
	panel = fill_gas_share_with_interpolation_and_fit(panel)
	panel = fill_nonfossil_share_with_interpolation_and_fit(panel)

	# Keep TCE decomposition fully populated for all years after share imputation.
	panel["CoalTCE"] = pd.to_numeric(panel["Energy"], errors="coerce") * pd.to_numeric(panel["CoalShare"], errors="coerce") / 100.0
	panel["OilTCE"] = pd.to_numeric(panel["Energy"], errors="coerce") * pd.to_numeric(panel["OilShare"], errors="coerce") / 100.0
	panel["GasTCE"] = pd.to_numeric(panel["Energy"], errors="coerce") * pd.to_numeric(panel["GasShare"], errors="coerce") / 100.0
	panel["NonFossilTCE"] = pd.to_numeric(panel["Energy"], errors="coerce") * pd.to_numeric(panel["NonFossilShare"], errors="coerce") / 100.0

	log_progress("阶段7/7：补全交通变量（HighwayMileage, PrivateCars）")
	prov_transport = read_provincial_transport_mileage_and_private_cars(items)
	log_observed_coverage("HighwayMileage", prov_transport, "HighwayMileage")
	log_observed_coverage("PrivateCars", prov_transport, "PrivateCars")
	if not prov_transport.empty:
		panel = panel.merge(prov_transport, on=["province", "year"], how="left")
	panel = ensure_metric_columns(panel, "HighwayMileage", include_imputed=True)
	panel = ensure_metric_columns(panel, "PrivateCars", include_imputed=True)

	panel = fill_positive_with_provincial_only(
		panel,
		value_col="HighwayMileage",
		interpolation_source="Provincial_highway_mileage_log_interpolation",
		polyfit_source="Provincial_highway_mileage_log_polynomial_fit",
	)
	panel = fill_positive_with_provincial_only(
		panel,
		value_col="PrivateCars",
		interpolation_source="Provincial_private_cars_log_interpolation",
		polyfit_source="Provincial_private_cars_log_polynomial_fit",
	)

	panel["A"] = panel["GDP"] / panel["Population"]
	panel["B"] = panel["Energy"] / panel["GDP"]
	panel["C"] = panel["CO2"] / panel["Energy"]
	panel["EnergyIntensity"] = panel["Energy"] / panel["GDP"]

	core_cols = [
    "province", "year", "CO2", "GDP", "Population", "Energy",
    "CoalTCE", "OilTCE", "GasTCE", "NonFossilTCE",
    "CoalShare", "OilShare", "GasShare", "NonFossilShare",
    "Industry", "Urbanization", "HighwayMileage", "PrivateCars",
    "A", "B", "C", "EnergyIntensity",   # ← 加上它
    "CO2_source",
	]
	panel_core = cast(pd.DataFrame, panel.loc[:, [c for c in core_cols if c in panel.columns]].copy())

	# Keep rows where CO2 exists; remaining variables are filled by available controls.
	panel_core = sort_panel_by_province_year(panel_core)
	lmdi_df = compute_lmdi_time(panel_core)
	national_energy_validation_df, national_energy_validation_stats = build_national_energy_validation(panel_core)
	lmdi_residual_ratio_stats = {}
	if not lmdi_df.empty and "lmdi_residual_abs_ratio" in lmdi_df.columns:
		lmdi_residual_ratio_stats = {
			"mean": float(lmdi_df["lmdi_residual_abs_ratio"].mean()),
			"median": float(lmdi_df["lmdi_residual_abs_ratio"].median()),
			"p90": float(lmdi_df["lmdi_residual_abs_ratio"].quantile(0.9)),
			"max": float(lmdi_df["lmdi_residual_abs_ratio"].max()),
		}

	log_progress("写出结果文件")

	panel_path = OUT_DIR / "panel_master.csv"
	lmdi_path = OUT_DIR / "lmdi_decomposition.csv"
	audit_path = OUT_DIR / "panel_source_audit.csv"
	summary_path = OUT_DIR / "panel_build_summary.json"
	national_energy_validation_path = OUT_DIR / "national_energy_validation.csv"

	panel_core.to_csv(panel_path, index=False, encoding="utf-8-sig")
	lmdi_df.to_csv(lmdi_path, index=False, encoding="utf-8-sig")
	national_energy_validation_df.to_csv(national_energy_validation_path, index=False, encoding="utf-8-sig")

	audit_cols = [
		c
		for c in panel.columns
		if c.endswith("_source") or c.endswith("_is_imputed")
	]
	panel[["province", "year", "CO2"] + audit_cols].to_csv(audit_path, index=False, encoding="utf-8-sig")

	summary = {
		"panel_rows": int(len(panel_core)),
		"province_count": int(panel_core["province"].nunique()),
		"year_min": int(panel_core["year"].min()) if len(panel_core) else None,
		"year_max": int(panel_core["year"].max()) if len(panel_core) else None,
		"co2_source_counts": panel_core["CO2_source"].value_counts(dropna=False).to_dict(),
		"lmdi_residual_abs_ratio_stats": lmdi_residual_ratio_stats,
		"missing_ratio": {
			c: float(panel_core[c].isna().mean())
			for c in [
				"CO2",
				"GDP",
				"Population",
				"Energy",
				"CoalTCE",
				"OilTCE",
				"GasTCE",
				"NonFossilTCE",
				"CoalShare",
				"OilShare",
				"GasShare",
				"NonFossilShare",
				"Industry",
				"Urbanization",
				"HighwayMileage",
				"PrivateCars",
				"A",
				"B",
				"C",
			]
			if c in panel_core.columns
		},
		"gdp_source_counts": panel["GDP_source"].value_counts(dropna=False).to_dict() if "GDP_source" in panel.columns else {},
		"population_source_counts": panel["Population_source"].value_counts(dropna=False).to_dict()
		if "Population_source" in panel.columns
		else {},
		"energy_source_counts": panel["Energy_source"].value_counts(dropna=False).to_dict()
		if "Energy_source" in panel.columns
		else {},
		"coal_share_source_counts": panel["CoalShare_source"].value_counts(dropna=False).to_dict()
		if "CoalShare_source" in panel.columns
		else {},
		"oil_share_source_counts": panel["OilShare_source"].value_counts(dropna=False).to_dict()
		if "OilShare_source" in panel.columns
		else {},
		"gas_share_source_counts": panel["GasShare_source"].value_counts(dropna=False).to_dict()
		if "GasShare_source" in panel.columns
		else {},
		"nonfossil_share_source_counts": panel["NonFossilShare_source"].value_counts(dropna=False).to_dict()
		if "NonFossilShare_source" in panel.columns
		else {},
		"industry_source_counts": panel["Industry_source"].value_counts(dropna=False).to_dict()
		if "Industry_source" in panel.columns
		else {},
		"urbanization_source_counts": panel["Urbanization_source"].value_counts(dropna=False).to_dict()
		if "Urbanization_source" in panel.columns
		else {},
		"highway_mileage_source_counts": panel["HighwayMileage_source"].value_counts(dropna=False).to_dict()
		if "HighwayMileage_source" in panel.columns
		else {},
		"private_cars_source_counts": panel["PrivateCars_source"].value_counts(dropna=False).to_dict()
		if "PrivateCars_source" in panel.columns
		else {},
		"national_energy_validation_stats": national_energy_validation_stats,
		"notes": [
			"CO2 uses MEIC provincial total emissions only (1990-2023).",
			"LMDI uses the four-factor Kaya identity: CO2 = Population * A * B * C.",
			"Energy is aggregated after converting fuel-specific total final consumption into standard coal equivalent (tce).",
			"Coal/Oil/Gas/NonFossil shares are defined as grouped tce divided by total final consumption tce.",
			"National energy total series is used for macro validation only and is not used as provincial fill value.",
			"All non-CO2 variables use provincial data only: in-province interpolation for internal gaps and polynomial fitting for leading/trailing gaps.",
			"No national fallback is used in this build.",
			"HighwayMileage and PrivateCars are integrated from provincial transport dataset.",
			"Panel rows are always output in stable province-year order.",
		],
	}
	summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

	log_progress("预处理流程完成")
	print("WROTE", panel_path)
	print("WROTE", lmdi_path)
	print("WROTE", audit_path)
	print("WROTE", national_energy_validation_path)
	print("WROTE", summary_path)
	print("PANEL_ROWS", summary["panel_rows"], "PROVINCES", summary["province_count"], "YEARS", summary["year_min"], summary["year_max"])
	print("CO2_SOURCE_COUNTS", summary["co2_source_counts"])


if __name__ == "__main__":
	main()
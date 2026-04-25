from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import torch
from openpyxl import load_workbook

# Ensure imports work when running as: python Code/Visual/compare_ipcc_vs_hybrid_2022.py
CODE_ROOT = Path(__file__).resolve().parents[1]
if str(CODE_ROOT) not in sys.path:
    sys.path.insert(0, str(CODE_ROOT))

from Model.stirpat_ee_gru import EntityEmbeddingGRU
from Preprocess.preprocess import ENERGY_TCE_FACTOR, normalize_province_name


COAL_FUEL_HEADERS = {
    "Raw_Coal",
    "Cleaned_Coal",
    "Other_Washed_Coal",
    "Briquettes",
    "Coke",
    "Coke_Oven_Gas",
    "Other_Gas",
    "Other_Coking_Products",
}

OIL_FUEL_HEADERS = {
    "Crude_Oil",
    "Gasoline",
    "Kerosene",
    "Diesel_Oil",
    "Fuel_Oil",
    "LPG",
    "Refinery_Gas",
    "Other_Petroleum_Products",
}

GAS_FUEL_HEADERS = {"Natural_Gas"}

NON_FOSSIL_HEADERS = {
    "Heat",
    "Electricity",
    "Other_Energy",
}

# IPCC coefficients from user-provided table.
IPCC_COEF = {
    "coal": 0.748,
    "oil": 0.583,
    "gas": 0.444,
    "non_fossil": 0.0,
}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compare IPCC coefficient method vs Hybrid(STIRPAT+DL) for provincial CO2 in 2022."
    )
    parser.add_argument("--year", type=int, default=2022, help="Target year")
    parser.add_argument(
        "--input-csv",
        default=str(CODE_ROOT / "Dataset" / "output" / "panel_with_residual.csv"),
        help="Path to panel_with_residual.csv",
    )
    parser.add_argument(
        "--dataset-npz",
        default=str(CODE_ROOT / "Dataset" / "output" / "stirpat_ee_gru_dataset.npz"),
        help="Dataset npz path",
    )
    parser.add_argument(
        "--model-ckpt",
        default=str(CODE_ROOT / "Model" / "output" / "best_ee_gru.pt"),
        help="Trained model checkpoint path",
    )
    parser.add_argument(
        "--energy-inventory-xlsx",
        default=str(
            CODE_ROOT.parents[0]
            / "Dataset"
            / "省份维度"
            / "能源相关"
            / "1997-2022分行业省级能源清单"
            / "省级能源清单_2022.xlsx"
        ),
        help="Provincial energy inventory file for target year",
    )
    parser.add_argument(
        "--output-dir",
        default=str(CODE_ROOT / "Visual" / "output"),
        help="Directory for outputs",
    )
    parser.add_argument(
        "--ipcc-energy-row",
        default="total_consumption",
        choices=["total_consumption", "total_final_consumption"],
        help="Energy inventory row used by IPCC method",
    )
    parser.add_argument("--device", default="auto", choices=["auto", "cpu", "cuda"], help="Inference device")
    return parser


def choose_device(device_arg: str) -> torch.device:
    if device_arg == "cpu":
        return torch.device("cpu")
    if device_arg == "cuda":
        if not torch.cuda.is_available():
            raise RuntimeError("CUDA requested but not available.")
        return torch.device("cuda")
    return torch.device("cuda" if torch.cuda.is_available() else "cpu")


def load_model(model_ckpt: Path, device: torch.device) -> EntityEmbeddingGRU:
    ckpt = torch.load(model_ckpt, map_location=device, weights_only=False)
    shape_info = ckpt["shape_info"]
    train_cfg = ckpt.get("train_config", {})

    model = EntityEmbeddingGRU(
        num_provinces=int(shape_info["num_provinces"]),
        num_dynamic_features=int(shape_info["num_dynamic_features"]),
        embed_dim=int(train_cfg.get("embed_dim", 8)),
        hidden_dim=int(train_cfg.get("hidden_dim", 32)),
        dropout=float(train_cfg.get("dropout", 0.2)),
    ).to(device)
    model.load_state_dict(ckpt["model_state_dict"])
    model.eval()
    return model


def infer_province_hybrid(
    province_df: pd.DataFrame,
    gru_feature_names: list[str],
    mean: np.ndarray,
    std: np.ndarray,
    window: int,
    model: EntityEmbeddingGRU,
    device: torch.device,
) -> pd.DataFrame:
    province_df = province_df.sort_values("year", kind="stable").reset_index(drop=True)
    province_df["year"] = province_df["year"].astype(int)
    province_df["true_co2"] = province_df["CO2"].astype(float)

    dynamic_raw = province_df[gru_feature_names].to_numpy(dtype=np.float32)
    dynamic_std = (dynamic_raw - mean) / std
    stirpat_log_arr = province_df["stirpat_log_pred"].to_numpy(dtype=np.float32)

    province_ids = province_df["province_id"].astype(int).unique().tolist()
    if len(province_ids) != 1:
        raise ValueError(f"Province has inconsistent province_id values: {province_ids}")
    province_idx = int(province_ids[0])

    pred_residual = np.zeros(len(province_df), dtype=np.float32)
    pred_log_co2 = np.zeros(len(province_df), dtype=np.float32)

    with torch.no_grad():
        for i in range(len(province_df)):
            dyn_seq = np.zeros((window, len(gru_feature_names)), dtype=np.float32)
            lag_seq = np.zeros((window, 1), dtype=np.float32)

            for j in range(window):
                src_idx = i - window + j
                if src_idx < 0:
                    dyn_seq[j] = dynamic_std[0]
                    lag_seq[j, 0] = 0.0
                else:
                    dyn_seq[j] = dynamic_std[src_idx]
                    lag_seq[j, 0] = pred_residual[src_idx]

            x_dyn = torch.tensor(dyn_seq[None, :, :], dtype=torch.float32, device=device)
            x_lag = torch.tensor(lag_seq[None, :, :], dtype=torch.float32, device=device)
            x_pid = torch.tensor([province_idx], dtype=torch.long, device=device)

            res_hat = float(model(x_dyn, x_lag, x_pid).cpu().item())
            pred_residual[i] = res_hat
            pred_log_co2[i] = float(stirpat_log_arr[i]) + res_hat

    province_df["pred_log_co2"] = pred_log_co2
    province_df["hybrid_co2"] = np.exp(province_df["pred_log_co2"]).astype(float)
    return province_df


def read_ipcc_2022(energy_inventory_xlsx: Path, year: int, energy_row_mode: str) -> pd.DataFrame:
    if not energy_inventory_xlsx.exists():
        raise FileNotFoundError(f"Energy inventory file not found: {energy_inventory_xlsx}")

    year_from_file = re.search(r"(19\d{2}|20\d{2})", energy_inventory_xlsx.name)
    if not year_from_file or int(year_from_file.group(1)) != year:
        raise ValueError(f"Energy inventory file year does not match --year={year}: {energy_inventory_xlsx.name}")

    wb = load_workbook(energy_inventory_xlsx, read_only=True, data_only=True)
    rows: list[dict[str, float | str]] = []

    def _calc_from_row(ws, row_idx: int) -> dict[str, float]:
        coal_tce = 0.0
        oil_tce = 0.0
        gas_tce = 0.0
        non_fossil_tce = 0.0

        max_col = int(ws.max_column or 0)
        for c in range(2, max_col + 1):
            header = ws.cell(row=1, column=c).value
            if not isinstance(header, str):
                continue
            header = header.strip()
            factor = ENERGY_TCE_FACTOR.get(header)
            if factor is None:
                continue

            v = ws.cell(row=row_idx, column=c).value
            if not isinstance(v, (int, float, np.integer, np.floating)):
                continue
            v_float = float(v)
            if not np.isfinite(v_float):
                continue

            contrib_tce = v_float * float(factor)
            if header in COAL_FUEL_HEADERS:
                coal_tce += contrib_tce
            elif header in OIL_FUEL_HEADERS:
                oil_tce += contrib_tce
            elif header in GAS_FUEL_HEADERS:
                gas_tce += contrib_tce
            elif header in NON_FOSSIL_HEADERS:
                non_fossil_tce += contrib_tce
            else:
                non_fossil_tce += contrib_tce

        total_tce = coal_tce + oil_tce + gas_tce + non_fossil_tce
        if total_tce <= 0:
            return {
                "ok": 0.0,
                "ipcc_co2": np.nan,
                "ipcc_total_tce": np.nan,
                "ipcc_coal_share_pct": np.nan,
                "ipcc_oil_share_pct": np.nan,
                "ipcc_gas_share_pct": np.nan,
                "ipcc_non_fossil_share_pct": np.nan,
            }

        coal_share = 100.0 * coal_tce / total_tce
        oil_share = 100.0 * oil_tce / total_tce
        gas_share = 100.0 * gas_tce / total_tce
        non_share = 100.0 * non_fossil_tce / total_tce

        ipcc_co2_mt = (
            (coal_tce * IPCC_COEF["coal"] + oil_tce * IPCC_COEF["oil"] + gas_tce * IPCC_COEF["gas"])
            * (44.0 / 12.0)
            / 100.0
        )

        return {
            "ok": 1.0,
            "ipcc_co2": float(ipcc_co2_mt),
            "ipcc_total_tce": float(total_tce),
            "ipcc_coal_share_pct": float(coal_share),
            "ipcc_oil_share_pct": float(oil_share),
            "ipcc_gas_share_pct": float(gas_share),
            "ipcc_non_fossil_share_pct": float(non_share),
        }

    def _is_nonphysical_shares(payload: dict[str, float]) -> bool:
        shares = [
            payload.get("ipcc_coal_share_pct", np.nan),
            payload.get("ipcc_oil_share_pct", np.nan),
            payload.get("ipcc_gas_share_pct", np.nan),
            payload.get("ipcc_non_fossil_share_pct", np.nan),
        ]
        for s in shares:
            if not np.isfinite(s):
                return True
            if s < -1e-6 or s > 100.0 + 1e-6:
                return True
        return False

    try:
        for sheet_name in wb.sheetnames:
            if str(sheet_name).upper() == "NOTE":
                continue

            province_key = re.sub(r"(19\d{2}|20\d{2})", "", str(sheet_name)).strip()
            province = normalize_province_name(province_key)
            if not province:
                continue

            ws = wb[sheet_name]
            max_row = int(ws.max_row or 0)
            max_col = int(ws.max_column or 0)
            if max_row < 3 or max_col < 2:
                continue

            row_map: dict[str, int] = {}
            for ridx in range(1, min(max_row, 120) + 1):
                v = ws.cell(row=ridx, column=1).value
                if isinstance(v, str) and v.strip():
                    row_map[v.strip().lower()] = ridx

            row_total = row_map.get("total consumption")
            row_final = row_map.get("total final consumption")
            selected_row = None
            row_label = None
            fallback_flag = 0
            fallback_reason = ""
            if energy_row_mode == "total_consumption":
                if row_total is not None:
                    payload_total = _calc_from_row(ws, row_total)
                    if payload_total.get("ok", 0.0) < 0.5:
                        selected_row = row_final
                        row_label = "Total Final Consumption"
                        fallback_flag = 1
                        fallback_reason = "total_consumption_invalid"
                        payload = _calc_from_row(ws, row_final) if row_final is not None else payload_total
                    elif _is_nonphysical_shares(payload_total):
                        # Total Consumption may include negative balance terms for some provinces.
                        # Fall back to Total Final Consumption to avoid non-physical shares.
                        selected_row = row_final
                        row_label = "Total Final Consumption"
                        fallback_flag = 1
                        fallback_reason = "nonphysical_shares_from_total_consumption"
                        payload = _calc_from_row(ws, row_final) if row_final is not None else payload_total
                    else:
                        selected_row = row_total
                        row_label = "Total Consumption"
                        payload = payload_total
                else:
                    selected_row = row_final
                    row_label = "Total Final Consumption"
                    fallback_flag = 1
                    fallback_reason = "total_consumption_missing"
                    payload = _calc_from_row(ws, row_final) if row_final is not None else {"ok": 0.0}
            else:
                if row_final is not None:
                    selected_row = row_final
                    row_label = "Total Final Consumption"
                    payload = _calc_from_row(ws, row_final)
                else:
                    selected_row = row_total
                    row_label = "Total Consumption"
                    fallback_flag = 1
                    fallback_reason = "total_final_consumption_missing"
                    payload = _calc_from_row(ws, row_total) if row_total is not None else {"ok": 0.0}

            if selected_row is None or payload.get("ok", 0.0) < 0.5:
                continue

            rows.append(
                {
                    "province": province,
                    "ipcc_co2": float(payload["ipcc_co2"]),
                    "ipcc_energy_row": str(row_label),
                    "ipcc_total_tce": float(payload["ipcc_total_tce"]),
                    "ipcc_coal_share_pct": float(payload["ipcc_coal_share_pct"]),
                    "ipcc_oil_share_pct": float(payload["ipcc_oil_share_pct"]),
                    "ipcc_gas_share_pct": float(payload["ipcc_gas_share_pct"]),
                    "ipcc_non_fossil_share_pct": float(payload["ipcc_non_fossil_share_pct"]),
                    "ipcc_row_fallback": int(fallback_flag),
                    "ipcc_row_fallback_reason": str(fallback_reason),
                }
            )
    finally:
        wb.close()

    out = pd.DataFrame(rows)
    if out.empty:
        raise RuntimeError("No valid province records parsed from energy inventory.")
    return out.sort_values("province", kind="stable").drop_duplicates(["province"], keep="last")


def calc_metrics(y_true: np.ndarray, y_pred: np.ndarray) -> dict[str, float]:
    err = y_pred - y_true
    mae = float(np.mean(np.abs(err)))
    rmse = float(np.sqrt(np.mean(err ** 2)))
    mape = float(np.mean(np.abs(err) / np.maximum(np.abs(y_true), 1e-6)) * 100.0)
    return {"mae": mae, "rmse": rmse, "mape_pct": mape}


def plot_comparison(compare_df: pd.DataFrame, metrics: dict[str, dict[str, float]], out_path: Path, year: int) -> None:
    plt.rcParams.update(
        {
            "font.size": 11,
            "axes.titlesize": 14,
            "axes.labelsize": 12,
            "xtick.labelsize": 10,
            "ytick.labelsize": 10,
            "legend.fontsize": 10,
        }
    )
    compare_df = compare_df.sort_values("actual_co2", ascending=False, kind="stable").reset_index(drop=True)
    y = np.arange(len(compare_df))

    fig = plt.figure(figsize=(22, 12), constrained_layout=True)
    gs = fig.add_gridspec(2, 2, width_ratios=[1.58, 1.0], height_ratios=[1.0, 1.0], wspace=0.06, hspace=0.12)

    ax0 = fig.add_subplot(gs[:, 0])
    ax1 = fig.add_subplot(gs[0, 1])
    ax2 = fig.add_subplot(gs[1, 1])

    for i in range(len(compare_df)):
        ax0.plot(
            [compare_df.loc[i, "actual_co2"], compare_df.loc[i, "hybrid_co2"]],
            [i, i],
            color="#1f77b4",
            alpha=0.30,
            linewidth=1.3,
        )
        ax0.plot(
            [compare_df.loc[i, "actual_co2"], compare_df.loc[i, "ipcc_co2"]],
            [i, i],
            color="#ff7f0e",
            alpha=0.30,
            linewidth=1.3,
        )

    ax0.scatter(compare_df["actual_co2"], y, color="#222222", s=38, label="Actual CO2", zorder=3)
    ax0.scatter(compare_df["hybrid_co2"], y, color="#1f77b4", s=34, label="Hybrid CO2", zorder=3)
    ax0.scatter(compare_df["ipcc_co2"], y, color="#ff7f0e", s=40, label="IPCC CO2", zorder=3)
    ax0.set_yticks(y)
    ax0.set_yticklabels(compare_df["province"], fontsize=10)
    ax0.invert_yaxis()
    ax0.grid(axis="x", linestyle="--", alpha=0.25)
    ax0.set_xlabel("CO2")
    ax0.set_title("Province-level CO2 Comparison", pad=12)
    ax0.legend(loc="lower right")

    width = 0.38
    ax1.bar(y - width / 2, compare_df["hybrid_ape_pct"], width=width, color="#1f77b4", alpha=0.85, label="Hybrid APE%")
    ax1.bar(y + width / 2, compare_df["ipcc_ape_pct"], width=width, color="#ff7f0e", alpha=0.80, label="IPCC APE%")
    ax1.set_xticks(y)
    ax1.set_xticklabels(compare_df["province"], rotation=75, ha="right", fontsize=8)
    ax1.set_ylabel("Absolute Percentage Error (%)")
    ax1.set_title("Province-level Absolute Percentage Error")
    ax1.grid(axis="y", linestyle="--", alpha=0.25)
    ax1.legend(loc="upper right")

    max_v = float(max(compare_df[["actual_co2", "hybrid_co2", "ipcc_co2"]].max()))
    min_v = float(min(compare_df[["actual_co2", "hybrid_co2", "ipcc_co2"]].min()))
    pad = (max_v - min_v) * 0.08 if max_v > min_v else 1.0
    lo, hi = min_v - pad, max_v + pad

    ax2.scatter(compare_df["actual_co2"], compare_df["hybrid_co2"], color="#1f77b4", s=42, alpha=0.88, label="Hybrid")
    ax2.scatter(compare_df["actual_co2"], compare_df["ipcc_co2"], color="#ff7f0e", s=42, alpha=0.80, label="IPCC")
    ax2.plot([lo, hi], [lo, hi], color="#444444", linestyle="--", linewidth=1.2, label="y=x")
    ax2.set_xlim(lo, hi)
    ax2.set_ylim(lo, hi)
    ax2.set_xlabel("Actual CO2")
    ax2.set_ylabel("Predicted CO2")
    ax2.set_title("Actual vs Predicted Scatter")
    ax2.grid(True, linestyle="--", alpha=0.25)
    ax2.legend(loc="upper left")

    txt = (
        f"Hybrid: MAE={metrics['hybrid']['mae']:.2f}, RMSE={metrics['hybrid']['rmse']:.2f}, MAPE={metrics['hybrid']['mape_pct']:.2f}%\n"
        f"IPCC:   MAE={metrics['ipcc']['mae']:.2f}, RMSE={metrics['ipcc']['rmse']:.2f}, MAPE={metrics['ipcc']['mape_pct']:.2f}%"
    )
    ax2.text(
        0.03,
        0.03,
        txt,
        transform=ax2.transAxes,
        va="bottom",
        ha="left",
        fontsize=10,
        bbox=dict(facecolor="white", edgecolor="#bbbbbb", alpha=0.85),
    )

    fig.suptitle(
        f"{year} CO2 Estimation Benchmark: IPCC Coefficient Method vs STIRPAT+Deep Learning",
        fontsize=16,
        y=1.02,
    )
    fig.savefig(str(out_path), dpi=320)
    plt.close(fig)


def main() -> None:
    args = build_parser().parse_args()

    year = int(args.year)
    input_csv = Path(args.input_csv)
    dataset_npz = Path(args.dataset_npz)
    model_ckpt = Path(args.model_ckpt)
    energy_inventory_xlsx = Path(args.energy_inventory_xlsx)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    device = choose_device(args.device)

    for p in [input_csv, dataset_npz, model_ckpt, energy_inventory_xlsx]:
        if not p.exists():
            raise FileNotFoundError(f"Required file not found: {p}")

    panel_df = pd.read_csv(input_csv)
    npz_data = np.load(dataset_npz, allow_pickle=True)
    model = load_model(model_ckpt=model_ckpt, device=device)

    if "gru_feature_names" not in npz_data or "standardizer_mean" not in npz_data or "standardizer_std" not in npz_data:
        raise KeyError("NPZ missing required metadata arrays: gru_feature_names/standardizer_mean/standardizer_std")
    if "train_dynamic_x" not in npz_data:
        raise KeyError("NPZ missing train_dynamic_x; cannot infer GRU window length")

    gru_feature_names = [str(x) for x in npz_data["gru_feature_names"].tolist()]
    mean = npz_data["standardizer_mean"].astype(np.float32)
    std = npz_data["standardizer_std"].astype(np.float32)
    window = int(npz_data["train_dynamic_x"].shape[1])

    province_names = sorted(panel_df["province"].dropna().astype(str).unique().tolist())

    hybrid_rows: list[dict[str, float | str]] = []
    for province in province_names:
        sub = panel_df.loc[(panel_df["province"] == province) & (panel_df["year"] <= year)].copy()
        if sub.empty:
            continue

        pred_df = infer_province_hybrid(
            province_df=sub,
            gru_feature_names=gru_feature_names,
            mean=mean,
            std=std,
            window=window,
            model=model,
            device=device,
        )

        row_y = pred_df.loc[pred_df["year"] == year]
        if row_y.empty:
            continue

        hybrid_rows.append(
            {
                "province": province,
                "actual_co2": float(row_y.iloc[0]["true_co2"]),
                "hybrid_co2": float(row_y.iloc[0]["hybrid_co2"]),
            }
        )

    hybrid_df = pd.DataFrame(hybrid_rows)
    ipcc_df = read_ipcc_2022(
        energy_inventory_xlsx=energy_inventory_xlsx,
        year=year,
        energy_row_mode=args.ipcc_energy_row,
    )

    compare_df = hybrid_df.merge(ipcc_df, on="province", how="inner")
    if compare_df.empty:
        raise RuntimeError("No overlapping provinces between hybrid output and IPCC inventory.")

    compare_df["hybrid_abs_err"] = (compare_df["hybrid_co2"] - compare_df["actual_co2"]).abs()
    compare_df["ipcc_abs_err"] = (compare_df["ipcc_co2"] - compare_df["actual_co2"]).abs()
    compare_df["hybrid_ape_pct"] = compare_df["hybrid_abs_err"] / np.maximum(compare_df["actual_co2"].abs(), 1e-6) * 100.0
    compare_df["ipcc_ape_pct"] = compare_df["ipcc_abs_err"] / np.maximum(compare_df["actual_co2"].abs(), 1e-6) * 100.0

    metrics = {
        "hybrid": calc_metrics(compare_df["actual_co2"].to_numpy(), compare_df["hybrid_co2"].to_numpy()),
        "ipcc": calc_metrics(compare_df["actual_co2"].to_numpy(), compare_df["ipcc_co2"].to_numpy()),
    }

    out_csv = output_dir / f"compare_ipcc_vs_hybrid_{year}.csv"
    out_json = output_dir / f"compare_ipcc_vs_hybrid_{year}_metrics.json"
    out_png = output_dir / f"compare_ipcc_vs_hybrid_{year}.png"

    compare_df.sort_values("province", kind="stable").to_csv(out_csv, index=False, encoding="utf-8")
    out_json.write_text(json.dumps(metrics, ensure_ascii=False, indent=2), encoding="utf-8")
    plot_comparison(compare_df=compare_df, metrics=metrics, out_path=out_png, year=year)

    panel_n = len(province_names)
    ipcc_n = len(ipcc_df)
    overlap_n = len(compare_df)

    print(f"Saved compare table: {out_csv}")
    print(f"Saved metrics json: {out_json}")
    print(f"Saved figure:       {out_png}")
    print(f"IPCC energy row mode: {args.ipcc_energy_row}")
    print(f"Panel provinces={panel_n}, IPCC inventory provinces={ipcc_n}, overlap={overlap_n}")
    print("Note: current data provide 30 provinces; Tibet is not available in panel/inventory.")
    print(
        f"Hybrid MAE={metrics['hybrid']['mae']:.3f}, MAPE={metrics['hybrid']['mape_pct']:.2f}% | "
        f"IPCC MAE={metrics['ipcc']['mae']:.3f}, MAPE={metrics['ipcc']['mape_pct']:.2f}%"
    )


if __name__ == "__main__":
    main()

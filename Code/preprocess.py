from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, cast

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATASET_ROOT = PROJECT_ROOT / "Dataset"
SUMMARY_JSON = PROJECT_ROOT / "Code" / "dataset_index_summary.json"
OUT_DIR = PROJECT_ROOT / "Code" / "output"


# Canonical province names aligned to inventory/MEIC naming.
PROVINCE_ALIAS = {
	"beijing": "Beijing",
	"tianjin": "Tianjin",
	"hebei": "Hebei",
	"shanxi": "Shanxi",
	"innermongolia": "InnerMongolia",
	"liaoning": "Liaoning",
	"jilin": "Jilin",
	"heilongjiang": "Heilongjiang",
	"shanghai": "Shanghai",
	"jiangsu": "Jiangsu",
	"zhejiang": "Zhejiang",
	"anhui": "Anhui",
	"fujian": "Fujian",
	"jiangxi": "Jiangxi",
	"shandong": "Shandong",
	"henan": "Henan",
	"hubei": "Hubei",
	"hunan": "Hunan",
	"guangdong": "Guangdong",
	"guangxi": "Guangxi",
	"hainan": "Hainan",
	"chongqing": "Chongqing",
	"sichuan": "Sichuan",
	"guizhou": "Guizhou",
	"yunnan": "Yunnan",
	"shaanxi": "Shaanxi",
	"gansu": "Gansu",
	"qinghai": "Qinghai",
	"ningxia": "Ningxia",
	"xinjiang": "Xinjiang",
	"xizang": "Tibet",
	# Common variants
	"innermongoliaautonomousregion": "InnerMongolia",
	"innermongoliaregion": "InnerMongolia",
	"tibet": "Tibet",
}

ZH_PROVINCE_ALIAS = {
	"北京": "Beijing",
	"北京市": "Beijing",
	"天津": "Tianjin",
	"天津市": "Tianjin",
	"河北": "Hebei",
	"河北省": "Hebei",
	"山西": "Shanxi",
	"山西省": "Shanxi",
	"内蒙古": "InnerMongolia",
	"内蒙古自治区": "InnerMongolia",
	"辽宁": "Liaoning",
	"辽宁省": "Liaoning",
	"吉林": "Jilin",
	"吉林省": "Jilin",
	"黑龙江": "Heilongjiang",
	"黑龙江省": "Heilongjiang",
	"上海": "Shanghai",
	"上海市": "Shanghai",
	"江苏": "Jiangsu",
	"江苏省": "Jiangsu",
	"浙江": "Zhejiang",
	"浙江省": "Zhejiang",
	"安徽": "Anhui",
	"安徽省": "Anhui",
	"福建": "Fujian",
	"福建省": "Fujian",
	"江西": "Jiangxi",
	"江西省": "Jiangxi",
	"山东": "Shandong",
	"山东省": "Shandong",
	"河南": "Henan",
	"河南省": "Henan",
	"湖北": "Hubei",
	"湖北省": "Hubei",
	"湖南": "Hunan",
	"湖南省": "Hunan",
	"广东": "Guangdong",
	"广东省": "Guangdong",
	"广西": "Guangxi",
	"广西壮族自治区": "Guangxi",
	"海南": "Hainan",
	"海南省": "Hainan",
	"重庆": "Chongqing",
	"重庆市": "Chongqing",
	"四川": "Sichuan",
	"四川省": "Sichuan",
	"贵州": "Guizhou",
	"贵州省": "Guizhou",
	"云南": "Yunnan",
	"云南省": "Yunnan",
	"陕西": "Shaanxi",
	"陕西省": "Shaanxi",
	"甘肃": "Gansu",
	"甘肃省": "Gansu",
	"青海": "Qinghai",
	"青海省": "Qinghai",
	"宁夏": "Ningxia",
	"宁夏回族自治区": "Ningxia",
	"新疆": "Xinjiang",
	"新疆维吾尔自治区": "Xinjiang",
	"西藏": "Tibet",
	"西藏自治区": "Tibet",
}


@dataclass
class VariableSeries:
	name: str
	source: str
	data: pd.DataFrame


def load_dataset_index() -> List[dict]:
	payload = json.loads(SUMMARY_JSON.read_text(encoding="utf-8"))
	return payload["items"]


def ensure_output_dir() -> None:
	OUT_DIR.mkdir(parents=True, exist_ok=True)


def normalize_province_name(name: str) -> Optional[str]:
	if not isinstance(name, str):
		return None
	name = name.strip()
	if name in ZH_PROVINCE_ALIAS:
		return ZH_PROVINCE_ALIAS[name]
	compact = re.sub(r"[^A-Za-z]", "", name).lower()
	return PROVINCE_ALIAS.get(compact)


def extract_year_from_text(text: str) -> Optional[int]:
	if not isinstance(text, str):
		return None
	m = re.search(r"(19\d{2}|20\d{2})", text)
	return int(m.group(1)) if m else None


def parse_year_cell(v) -> Optional[int]:
	if isinstance(v, (int, float)):
		iv = int(v)
		if 1900 <= iv <= 2100:
			return iv
	elif isinstance(v, str):
		return extract_year_from_text(v)
	return None


def find_header_row(ws) -> int:
	max_row = ws.max_row or 0
	max_col = ws.max_column or 0
	scan_rows = min(max_row, 15)

	best_row = 1
	best_year_hits = -1

	# Primary strategy: row with the largest count of year-like columns is the header row.
	for ridx in range(1, scan_rows + 1):
		hits = 0
		for c in range(2, max_col + 1):
			if parse_year_cell(ws.cell(row=ridx, column=c).value) is not None:
				hits += 1
		if hits > best_year_hits:
			best_year_hits = hits
			best_row = ridx

	if best_year_hits >= 5:
		return best_row

	# Fallback strategy for non-standard templates.
	for ridx in range(1, min(max_row, 25) + 1):
		val = ws.cell(row=ridx, column=1).value
		if isinstance(val, str) and val.strip() == "指标":
			return ridx

	return 1


def parse_year_headers(ws, header_row: int) -> Dict[int, int]:
	year_to_col: Dict[int, int] = {}
	max_col = ws.max_column or 0
	for col in range(2, max_col + 1):
		v = ws.cell(row=header_row, column=col).value
		year = None
		year = parse_year_cell(v)
		if year is not None:
			year_to_col[year] = col
	return year_to_col


def to_float(v) -> Optional[float]:
	if v is None:
		return None
	if isinstance(v, (int, float, np.number)):
		return float(v)
	if isinstance(v, str):
		s = v.strip().replace(",", "").replace("%", "")
		if not s:
			return None
		try:
			return float(s)
		except ValueError:
			return None
	return None


def ensure_cols(df: pd.DataFrame, defaults: Dict[str, Any]) -> pd.DataFrame:
	for col, default_value in defaults.items():
		if col not in df.columns:
			df[col] = cast(Any, default_value)
	return df


def read_inventory_co2_nbs(items: List[dict]) -> pd.DataFrame:
	records: List[dict] = []

	inventory_files = sorted(
		[
			x["path"]
			for x in items
			if x.get("sheet_count") == 31 and "2001-2022" in x["path"]
		]
	)

	for path_str in inventory_files:
		year = extract_year_from_text(Path(path_str).name)
		if year is None:
			continue

		wb = load_workbook(path_str, read_only=True, data_only=True)
		for sheet_name in wb.sheetnames:
			if str(sheet_name).upper() == "NOTE":
				continue

			province = normalize_province_name(sheet_name)
			if not province:
				continue

			ws = wb[sheet_name]
			max_row = ws.max_row or 0
			max_col = ws.max_column or 0
			if max_row < 4 or max_col < 5:
				continue

			header = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
			header_norm = [str(h).strip() if h is not None else "" for h in header]

			try:
				total_row = None
				for r in range(2, max_row + 1):
					v = ws.cell(row=r, column=1).value
					if isinstance(v, str) and v.strip() == "TotalEmissions":
						total_row = r
						break
				if total_row is None:
					continue

				if "Scope_1_Total" in header_norm:
					cidx = header_norm.index("Scope_1_Total") + 1
					co2 = to_float(ws.cell(row=total_row, column=cidx).value)
				else:
					# Fallback: sum known fuel/process columns in the total row.
					cols = [
						i + 1
						for i, h in enumerate(header_norm)
						if h
						and h
						not in {
							"Emission_Inventory",
							"unit",
						}
						and not h.startswith("Scope_2")
					]
					vals = [to_float(ws.cell(row=total_row, column=c).value) for c in cols]
					co2 = float(np.nansum([v for v in vals if v is not None])) if vals else None

				if co2 is None:
					continue

				records.append(
					{
						"province": province,
						"year": int(year),
						"CO2": float(co2),
						"CO2_source": "NBS_2001_2022_inventory",
						"CO2_source_priority": 1,
					}
				)
			except Exception:
				continue

	return pd.DataFrame(records)


def read_meic_co2(items: List[dict]) -> pd.DataFrame:
	meic_path = next((x["path"] for x in items if "MEIC" in x["path"]), None)
	if not meic_path:
		return pd.DataFrame(columns=["province", "year", "CO2", "CO2_source", "CO2_source_priority"])

	raw = pd.read_excel(meic_path, sheet_name="MEIC-China-CO2 total emissions", header=8)
	raw = raw.rename(columns={"Province": "province_raw", "Sector": "sector"})
	raw = raw[raw["sector"].astype(str).str.lower() == "total"].copy()

	year_cols = [c for c in raw.columns if isinstance(c, (int, float)) and 1900 <= int(c) <= 2100]
	if not year_cols:
		return pd.DataFrame(columns=["province", "year", "CO2", "CO2_source", "CO2_source_priority"])

	long_df = raw.melt(
		id_vars=["province_raw"],
		value_vars=year_cols,
		var_name="year",
		value_name="CO2",
	)
	long_df["year"] = long_df["year"].astype(int)
	long_df = long_df[(long_df["year"] >= 1990) & (long_df["year"] <= 2022)]
	long_df["province"] = long_df["province_raw"].map(normalize_province_name)
	long_df["CO2"] = pd.to_numeric(long_df["CO2"], errors="coerce")
	long_df = long_df.dropna(subset=["province", "CO2"])

	out = long_df[["province", "year", "CO2"]].copy()
	out["CO2_source"] = "MEIC_1990_2022"
	out["CO2_source_priority"] = 1
	return out


def align_meic_to_nbs_scale(nbs: pd.DataFrame, meic: pd.DataFrame) -> pd.DataFrame:
	if nbs.empty or meic.empty:
		return meic

	overlap_years = [1998, 1999, 2000]
	nbs_anchor = nbs[(nbs["year"] >= 2001) & (nbs["year"] <= 2003)]
	meic_anchor = meic[meic["year"].isin(overlap_years)]
	if nbs_anchor.empty or meic_anchor.empty:
		return meic

	nbs_mean = float(nbs_anchor["CO2"].mean())
	meic_mean = float(meic_anchor["CO2"].mean())
	global_scale = nbs_mean / meic_mean if np.isfinite(nbs_mean) and np.isfinite(meic_mean) and meic_mean > 0 else 1.0
	if not np.isfinite(global_scale) or global_scale <= 0:
		global_scale = 1.0

	aligned = meic.copy()
	aligned["CO2_scale_factor"] = global_scale
	aligned["CO2"] = aligned["CO2"] * global_scale
	aligned["CO2_source"] = aligned["CO2_source"].astype(str) + "_scaled_to_NBS_anchor_global"
	return aligned


def build_co2_panel(items: List[dict]) -> pd.DataFrame:
	meic = read_meic_co2(items)
	meic = meic[(meic["year"] >= 1990) & (meic["year"] <= 2022)].copy()
	meic = meic.sort_values(["province", "year", "CO2_source_priority"]).drop_duplicates(
		subset=["province", "year"],
		keep="first",
	)
	return meic.sort_values(["province", "year"]).reset_index(drop=True)


def read_provincial_energy_inventory(items: List[dict]) -> pd.DataFrame:
	records: List[dict] = []

	energy_files = sorted(
		[
			x["path"]
			for x in items
			if Path(x["path"]).name.startswith("省级能源清单_") and Path(x["path"]).suffix.lower() == ".xlsx"
		]
	)

	for path_str in energy_files:
		year = extract_year_from_text(Path(path_str).name)
		if year is None or year < 1990 or year > 2022:
			continue

		wb = load_workbook(path_str, read_only=True, data_only=True)
		for sheet_name in wb.sheetnames:
			if str(sheet_name).upper() == "NOTE":
				continue

			province = normalize_province_name(sheet_name)
			if not province:
				continue

			ws = wb[sheet_name]
			max_row = int(ws.max_row or 0)
			max_col = int(ws.max_column or 0)
			if max_row < 3 or max_col < 2:
				continue

			total_row = None
			for ridx in range(1, min(max_row, 120) + 1):
				v = ws.cell(row=ridx, column=1).value
				if isinstance(v, str) and v.strip().lower() == "total final consumption":
					total_row = ridx
					break

			if total_row is None:
				continue

			vals = [to_float(ws.cell(row=total_row, column=c).value) for c in range(2, max_col + 1)]
			nums = [float(v) for v in vals if v is not None and np.isfinite(v)]
			if not nums:
				continue

			records.append(
				{
					"province": province,
					"year": int(year),
					"Energy": float(np.nansum(nums)),
					"Energy_source": "Provincial_energy_inventory_1997_2022_total_final_consumption_sum",
					"Energy_is_national_proxy": 0,
				}
			)

	df = pd.DataFrame(records)
	if df.empty:
		return pd.DataFrame(columns=["province", "year", "Energy", "Energy_source", "Energy_is_national_proxy"])

	df = df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last")
	return df.reset_index(drop=True)


def select_single_sheet_path(items: List[dict], keyword: str) -> Optional[str]:
	for it in items:
		if it.get("sheet_count") == 1 and keyword in it["path"]:
			return it["path"]
	return None


def select_single_sheet_path_by_keywords(items: List[dict], keywords: Iterable[str]) -> Optional[str]:
	for it in items:
		if it.get("sheet_count") != 1:
			continue
		path = str(it.get("path", ""))
		if all(k in path for k in keywords):
			return path
	return None


def read_national_series(
	path_str: str,
	variable_name: str,
	source_name: str,
	indicator_keywords: Iterable[str],
) -> VariableSeries:
	wb = load_workbook(path_str, read_only=True, data_only=True)
	ws = wb[wb.sheetnames[0]]
	header_row = find_header_row(ws)
	year_cols = parse_year_headers(ws, header_row)

	target_row = None
	max_row = ws.max_row or 0
	candidates: List[tuple] = []
	for ridx in range(header_row + 1, max_row + 1):
		label = ws.cell(row=ridx, column=1).value
		if not isinstance(label, str):
			continue
		label = label.strip()
		if not label:
			continue

		vals = [to_float(ws.cell(row=ridx, column=c).value) for c in year_cols.values()]
		non_null = sum(v is not None for v in vals)
		numeric_vals = [v for v in vals if v is not None]
		std = float(np.std(numeric_vals)) if len(numeric_vals) >= 2 else 0.0

		if any(k in label for k in indicator_keywords):
			# Rank by coverage first, then by variability to avoid selecting all-100 total rows.
			candidates.append((non_null, std, ridx))

	if candidates:
		candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
		target_row = candidates[0][2]

	if target_row is None:
		# Fallback: first non-empty numeric row after header.
		for ridx in range(header_row + 1, max_row + 1):
			label = ws.cell(row=ridx, column=1).value
			if not isinstance(label, str) or not label.strip():
				continue
			vals = [to_float(ws.cell(row=ridx, column=c).value) for c in year_cols.values()]
			if any(v is not None for v in vals):
				target_row = ridx
				break

	rows: List[dict] = []
	if target_row is not None:
		for year, col in sorted(year_cols.items()):
			rows.append({"year": int(year), variable_name: to_float(ws.cell(row=target_row, column=col).value)})

	df = pd.DataFrame(rows, columns=["year", variable_name])
	if not df.empty:
		df = cast(pd.DataFrame, df.loc[(df["year"] >= 1990) & (df["year"] <= 2022), :].copy())
	return VariableSeries(name=variable_name, source=source_name, data=cast(pd.DataFrame, df))


def build_national_controls(items: List[dict]) -> List[VariableSeries]:
	# National proxies remain needed for fallback and anchoring.
	gdp_path = select_single_sheet_path_by_keywords(items, ["国家维度", "经济相关", "人均GDP"])
	population_path = select_single_sheet_path_by_keywords(items, ["国家维度", "人口相关", "年末总人口"])
	urban_path = select_single_sheet_path_by_keywords(items, ["国家维度", "人口相关", "城镇化率"])
	industry_path = select_single_sheet_path_by_keywords(items, ["国家维度", "经济相关", "三次产业构成"])
	energy_path = select_single_sheet_path_by_keywords(items, ["国家维度", "能源相关", "能源消费总量"])

	if gdp_path is None:
		gdp_path = select_single_sheet_path(items, "GDP")

	series_list: List[VariableSeries] = []
	population_series: Optional[VariableSeries] = None
	urban_population_series: Optional[VariableSeries] = None
	urbanization_series: Optional[VariableSeries] = None
	if gdp_path:
		series_list.append(
			read_national_series(
				gdp_path,
				variable_name="GDP",
				source_name="National_macro_GDP",
				indicator_keywords=["国民总收入", "国内生产总值", "GDP"],
			)
		)
	if population_path:
		series_list.append(
			read_national_series(
				population_path,
				variable_name="Population",
				source_name="National_macro_Population",
				indicator_keywords=["年末总人口"],
			)
		)
		population_series = series_list[-1]
		# Prefer computing urbanization from population components when available.
		series_list.append(
			read_national_series(
				population_path,
				variable_name="UrbanPopulation",
				source_name="National_macro_UrbanPopulation",
				indicator_keywords=["城镇人口"],
			)
		)
		urban_population_series = series_list[-1]
	if energy_path:
		series_list.append(
			read_national_series(
				energy_path,
				variable_name="Energy",
				source_name="National_macro_Energy",
				indicator_keywords=["能源消费总量", "总量"],
			)
		)
	if industry_path:
		series_list.append(
			read_national_series(
				industry_path,
				variable_name="Industry",
				source_name="National_macro_IndustryShare",
				indicator_keywords=["第二产业"],
			)
		)
	if urban_path:
		urbanization_series = read_national_series(
			urban_path,
			variable_name="Urbanization",
			source_name="National_macro_Urbanization",
			indicator_keywords=["城镇化率", "城镇"],
		)
		if not urbanization_series.data.empty:
			series_list.append(urbanization_series)

	if (urbanization_series is None or urbanization_series.data.empty) and population_series is not None and urban_population_series is not None:
		fallback_df = population_series.data.merge(
			urban_population_series.data,
			on="year",
			how="inner",
			suffixes=("_Population", "_UrbanPopulation"),
		)
		if not fallback_df.empty and "Population" in fallback_df.columns and "UrbanPopulation" in fallback_df.columns:
			fallback_df["Urbanization"] = (pd.to_numeric(fallback_df["UrbanPopulation"], errors="coerce") / pd.to_numeric(fallback_df["Population"], errors="coerce")) * 100.0
			fallback_df = cast(pd.DataFrame, fallback_df.loc[:, ["year", "Urbanization"]].copy())
			fallback_df = cast(pd.DataFrame, fallback_df.dropna(subset=["Urbanization"]))
			if not fallback_df.empty:
				series_list.append(
					VariableSeries(
						name="Urbanization",
						source="National_macro_Urbanization_derived_from_UrbanPopulation_over_Population",
						data=fallback_df.sort_values("year").reset_index(drop=True),
					)
				)

	return series_list


def read_provincial_wide_series(
	items: List[dict],
	path_keywords: Iterable[str],
	variable_name: str,
	source_name: str,
) -> VariableSeries:
	path_str = select_single_sheet_path_by_keywords(items, path_keywords)
	if not path_str:
		return VariableSeries(name=variable_name, source=source_name, data=pd.DataFrame(columns=["province", "year", variable_name]))

	wb = load_workbook(path_str, read_only=True, data_only=True)
	ws = wb[wb.sheetnames[0]]
	year_cols: Dict[int, int] = {}
	for col in range(3, int(ws.max_column or 0) + 1):
		year = parse_year_cell(ws.cell(row=1, column=col).value)
		if year is not None and 1990 <= int(year) <= 2022:
			year_cols[int(year)] = col

	rows: List[dict] = []
	for ridx in range(2, int(ws.max_row or 0) + 1):
		province_raw = ws.cell(row=ridx, column=2).value
		province = normalize_province_name(str(province_raw))
		if not province:
			continue
		for year, col in sorted(year_cols.items()):
			v = to_float(ws.cell(row=ridx, column=col).value)
			if v is None:
				continue
			rows.append({"province": province, "year": int(year), variable_name: float(v)})

	df = pd.DataFrame(rows, columns=["province", "year", variable_name])
	if not df.empty:
		df = cast(pd.DataFrame, df.loc[df["year"].between(1990, 2022), :].copy())
		df = df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last").reset_index(drop=True)
	return VariableSeries(name=variable_name, source=source_name, data=cast(pd.DataFrame, df))


def merge_provincial_variable(panel: pd.DataFrame, series: VariableSeries) -> pd.DataFrame:
	if series.data.empty:
		return panel
	value_col = series.name
	temp_col = f"{value_col}_provincial_ref"
	df = panel.merge(series.data.rename(columns={value_col: temp_col}), on=["province", "year"], how="left")
	observed_mask = df[temp_col].notna()
	df.loc[observed_mask, value_col] = pd.to_numeric(df.loc[observed_mask, temp_col], errors="coerce")
	df.loc[observed_mask, f"{value_col}_source"] = series.source
	df.loc[observed_mask, f"{value_col}_is_national_proxy"] = 0
	df = df.drop(columns=[temp_col], errors="ignore")
	return df


def transformed_interpolate_extrapolate_by_province(
	df: pd.DataFrame,
	value_col: str,
	source_col: str,
	proxy_flag_col: str,
	imputed_col: str,
	mode: str,
	inside_source: str,
	fit_source: str,
	backcast_source: str,
) -> pd.DataFrame:
	data = ensure_cols(
		df.copy(),
		{
			value_col: np.nan,
			source_col: np.nan,
			proxy_flag_col: np.nan,
			imputed_col: 0,
		},
	)

	def _to_transformed(series: pd.Series) -> pd.Series:
		if mode == "log":
			base = pd.to_numeric(series, errors="coerce")
			vals = np.log(base.where(base > 0))
			return pd.Series(vals, index=series.index)
		if mode == "logit":
			base = pd.to_numeric(series, errors="coerce")
			base = base.where((base > 0) & (base < 100))
			base = base.clip(lower=0.01, upper=99.99)
			vals = np.log(base / (100.0 - base))
			return pd.Series(vals, index=series.index)
		raise ValueError(f"Unsupported mode: {mode}")

	def _from_transformed(series: pd.Series) -> pd.Series:
		if mode == "log":
			vals = np.exp(series)
			return pd.Series(vals, index=series.index)
		if mode == "logit":
			vals = 100.0 / (1.0 + np.exp(-series))
			vals_series = pd.Series(vals, index=series.index)
			return vals_series.clip(lower=0.01, upper=99.99)
		raise ValueError(f"Unsupported mode: {mode}")

	for province, idx in data.groupby("province").groups.items():
		sub = cast(pd.DataFrame, data.loc[list(idx), :].sort_values("year").copy())
		sub[value_col] = pd.to_numeric(sub[value_col], errors="coerce")
		if mode == "log":
			sub.loc[sub[value_col] <= 0, value_col] = np.nan
		else:
			sub.loc[(sub[value_col] <= 0) | (sub[value_col] >= 100), value_col] = np.nan

		transformed = _to_transformed(sub[value_col])
		transformed_interp = transformed.interpolate(method="linear", limit_area="inside")
		inside_mask = sub[value_col].isna() & transformed_interp.notna()
		if inside_mask.any():
			sub.loc[inside_mask, value_col] = _from_transformed(transformed_interp.loc[inside_mask])
			sub.loc[inside_mask, source_col] = inside_source
			sub.loc[inside_mask, proxy_flag_col] = 0
			sub.loc[inside_mask, imputed_col] = 1

		remain_mask = sub[value_col].isna()
		obs_mask = sub[value_col].notna()
		if remain_mask.any() and int(obs_mask.sum()) >= 4:
			x = sub.loc[obs_mask, "year"].astype(float).to_numpy()
			y = _to_transformed(sub.loc[obs_mask, value_col]).astype(float).to_numpy()
			xm = sub.loc[remain_mask, "year"].astype(float).to_numpy()
			degree = 2 if int(obs_mask.sum()) >= 6 else 1
			coeffs = np.polyfit(x, y, degree)
			pred = _from_transformed(pd.Series(np.polyval(coeffs, xm), index=sub.loc[remain_mask].index))
			sub.loc[remain_mask, value_col] = pred
			sub.loc[remain_mask, source_col] = fit_source if degree == 1 else fit_source + "_quadratic"
			sub.loc[remain_mask, proxy_flag_col] = 0
			sub.loc[remain_mask, imputed_col] = 1
		elif remain_mask.any() and int(obs_mask.sum()) >= 2:
			obs = cast(pd.DataFrame, sub.loc[obs_mask, ["year", value_col]].sort_values("year").copy())
			x1 = float(obs.iloc[0]["year"])
			x2 = float(obs.iloc[1]["year"])
			y1 = float(_to_transformed(pd.Series([obs.iloc[0][value_col]])).iloc[0])
			y2 = float(_to_transformed(pd.Series([obs.iloc[1][value_col]])).iloc[0])
			slope = (y2 - y1) / (x2 - x1) if abs(x2 - x1) > 0 else 0.0
			xm = sub.loc[remain_mask, "year"].astype(float).to_numpy()
			pred = _from_transformed(pd.Series(y1 + slope * (xm - x1), index=sub.loc[remain_mask].index))
			sub.loc[remain_mask, value_col] = pred
			sub.loc[remain_mask, source_col] = backcast_source
			sub.loc[remain_mask, proxy_flag_col] = 0
			sub.loc[remain_mask, imputed_col] = 1

		data.loc[sub.index, [value_col, source_col, proxy_flag_col, imputed_col]] = sub[
			[value_col, source_col, proxy_flag_col, imputed_col]
		]

	data[value_col] = pd.to_numeric(data[value_col], errors="coerce")
	data[proxy_flag_col] = pd.to_numeric(data[proxy_flag_col], errors="coerce").fillna(0)
	data[imputed_col] = pd.to_numeric(data[imputed_col], errors="coerce").fillna(0)
	return data


def apply_national_anchor_shift(
	df: pd.DataFrame,
	value_col: str,
	source_col: str,
	proxy_flag_col: str,
	imputed_col: str,
	national_col: str,
	fallback_source: str,
) -> pd.DataFrame:
	data = df.copy()
	for year, idx in data.groupby("year").groups.items():
		year_idx = list(idx)
		nat_vals = data.loc[year_idx, national_col].dropna()
		if nat_vals.empty:
			continue
		nat = float(nat_vals.iloc[0])
		year_mask = data.index.isin(year_idx)
		imputed_mask = year_mask & (data[imputed_col] == 1) & data[value_col].notna()
		if not imputed_mask.any():
			continue
		mean_year = float(data.loc[year_mask & data[value_col].notna(), value_col].mean())
		if not np.isfinite(mean_year):
			continue
		delta = nat - mean_year
		data.loc[imputed_mask, value_col] = data.loc[imputed_mask, value_col] + delta
		if value_col in {"Industry", "Urbanization"}:
			data.loc[imputed_mask, value_col] = data.loc[imputed_mask, value_col].clip(lower=0.01, upper=99.99)
		data.loc[imputed_mask, source_col] = data.loc[imputed_mask, source_col].astype(str) + "_national_anchor_shift"

	missing_mask = data[value_col].isna() & data[national_col].notna()
	data.loc[missing_mask, value_col] = data.loc[missing_mask, national_col]
	data.loc[missing_mask, source_col] = fallback_source
	data.loc[missing_mask, proxy_flag_col] = 1
	data.loc[missing_mask, imputed_col] = 0

	data = data.drop(columns=[national_col], errors="ignore")
	data[proxy_flag_col] = pd.to_numeric(data[proxy_flag_col], errors="coerce").fillna(0)
	data[imputed_col] = pd.to_numeric(data[imputed_col], errors="coerce").fillna(0)
	return data


def read_provincial_industry_share(items: List[dict]) -> pd.DataFrame:
	path_str = select_single_sheet_path_by_keywords(items, ["省份维度", "经济相关", "省份第二产业占比"])
	if not path_str:
		return pd.DataFrame(columns=["province", "year", "Industry", "Industry_source", "Industry_is_national_proxy"])

	wb = load_workbook(path_str, read_only=True, data_only=True)
	ws = wb[wb.sheetnames[0]]

	year_cols: Dict[int, int] = {}
	for col in range(3, int(ws.max_column or 0) + 1):
		year = parse_year_cell(ws.cell(row=1, column=col).value)
		if year is not None and 1990 <= int(year) <= 2022:
			year_cols[int(year)] = col

	if not year_cols:
		return pd.DataFrame(columns=["province", "year", "Industry", "Industry_source", "Industry_is_national_proxy"])

	records: List[dict] = []
	for ridx in range(2, int(ws.max_row or 0) + 1):
		indicator = ws.cell(row=ridx, column=1).value
		province_raw = ws.cell(row=ridx, column=2).value
		if not isinstance(indicator, str):
			continue
		if not isinstance(province_raw, str):
			continue
		if "第二产业" not in indicator or "比重" not in indicator:
			continue

		province = normalize_province_name(province_raw)
		if not province:
			continue

		for year, col in sorted(year_cols.items()):
			v = to_float(ws.cell(row=ridx, column=col).value)
			if v is None:
				continue
			records.append(
				{
					"province": province,
					"year": int(year),
					"Industry": float(v),
					"Industry_source": "Provincial_industry_share_1996_2024_observed",
					"Industry_is_national_proxy": 0,
				}
			)

	df = pd.DataFrame(records)
	if df.empty:
		return pd.DataFrame(columns=["province", "year", "Industry", "Industry_source", "Industry_is_national_proxy"])

	df = df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last")
	return df.reset_index(drop=True)


def fill_industry_with_interpolation_fit_and_anchor(
	panel: pd.DataFrame,
	industry_proxy_series: Optional[VariableSeries],
) -> pd.DataFrame:
	df = transformed_interpolate_extrapolate_by_province(
		panel,
		value_col="Industry",
		source_col="Industry_source",
		proxy_flag_col="Industry_is_national_proxy",
		imputed_col="Industry_is_imputed",
		mode="logit",
		inside_source="Provincial_industry_logit_interpolation",
		fit_source="Provincial_industry_logit_linear_fit_extrapolation",
		backcast_source="Provincial_industry_logit_backcast_extrapolation",
	)

	if industry_proxy_series is not None and not industry_proxy_series.data.empty:
		proxy = industry_proxy_series.data.rename(columns={"Industry": "Industry_national_ref"})
		df = df.merge(proxy, on="year", how="left")
		df = apply_national_anchor_shift(
			df,
			value_col="Industry",
			source_col="Industry_source",
			proxy_flag_col="Industry_is_national_proxy",
			imputed_col="Industry_is_imputed",
			national_col="Industry_national_ref",
			fallback_source=industry_proxy_series.source,
		)

	return df


def read_provincial_urbanization_share(items: List[dict]) -> pd.DataFrame:
	path_str = select_single_sheet_path_by_keywords(items, ["省份维度", "人口相关", "城镇人口所占比重"])
	if not path_str:
		return pd.DataFrame(columns=["province", "year", "Urbanization_prov"])

	wb = load_workbook(path_str, read_only=True, data_only=True)
	ws = wb[wb.sheetnames[0]]

	year_cols: Dict[int, int] = {}
	for col in range(3, int(ws.max_column or 0) + 1):
		year = parse_year_cell(ws.cell(row=1, column=col).value)
		if year is not None and 1990 <= int(year) <= 2022:
			year_cols[int(year)] = col

	if not year_cols:
		return pd.DataFrame(columns=["province", "year", "Urbanization_prov"])

	records: List[dict] = []
	for ridx in range(2, int(ws.max_row or 0) + 1):
		indicator = ws.cell(row=ridx, column=1).value
		province_raw = ws.cell(row=ridx, column=2).value
		if not isinstance(indicator, str):
			continue
		if not isinstance(province_raw, str):
			continue
		if "城镇人口" not in indicator or "比重" not in indicator:
			continue

		province = normalize_province_name(province_raw)
		if not province:
			continue

		for year, col in sorted(year_cols.items()):
			v = to_float(ws.cell(row=ridx, column=col).value)
			if v is None:
				continue
			records.append(
				{
					"province": province,
					"year": int(year),
					"Urbanization_prov": float(v),
				}
			)

	df = pd.DataFrame(records)
	if df.empty:
		return pd.DataFrame(columns=["province", "year", "Urbanization_prov"])

	df = df.sort_values(["province", "year"]).drop_duplicates(subset=["province", "year"], keep="last")
	return df.reset_index(drop=True)


def fill_urbanization_with_interpolation_fit_and_anchor(
	panel: pd.DataFrame,
	urbanization_proxy_series: Optional[VariableSeries],
) -> pd.DataFrame:
	df = ensure_cols(
		panel.copy(),
		{
			"Urbanization": np.nan,
			"Urbanization_source": np.nan,
			"Urbanization_is_national_proxy": np.nan,
			"Urbanization_is_imputed": 0,
		},
	)
	df["Urbanization_source"] = df["Urbanization_source"].astype(object)
	df["Urbanization_national_ref"] = pd.to_numeric(df["Urbanization"], errors="coerce") if "Urbanization" in df.columns else np.nan

	if "Urbanization_prov" in df.columns:
		obs_mask = df["Urbanization_prov"].notna()
		df.loc[obs_mask, "Urbanization"] = pd.to_numeric(df.loc[obs_mask, "Urbanization_prov"], errors="coerce")
		df.loc[obs_mask, "Urbanization_source"] = "Provincial_urbanization_share_observed"
		df.loc[obs_mask, "Urbanization_is_national_proxy"] = 0
		df.loc[obs_mask, "Urbanization_is_imputed"] = 0

	for province, idx in df.groupby("province").groups.items():
		province_idx = list(idx)
		province_mask = df.index.isin(province_idx)
		ref_mask = province_mask & df["Urbanization_national_ref"].notna()
		obs_proxy_mask = province_mask & df["Urbanization"].notna() & df["Urbanization_national_ref"].notna()
		alpha = np.nan
		if obs_proxy_mask.any():
			ratio = pd.to_numeric(df.loc[obs_proxy_mask, "Urbanization"], errors="coerce") / pd.to_numeric(df.loc[obs_proxy_mask, "Urbanization_national_ref"], errors="coerce")
			ratio = ratio.replace([np.inf, -np.inf], np.nan).dropna()
			if not ratio.empty:
				alpha = float(ratio.mean())
		if not np.isfinite(alpha) or alpha <= 0:
			alpha = 1.0
		fill_mask = province_mask & df["Urbanization_prov"].isna() if "Urbanization_prov" in df.columns else province_mask
		fill_mask = fill_mask & df["Urbanization_national_ref"].notna()
		if fill_mask.any():
			df.loc[fill_mask, "Urbanization"] = pd.to_numeric(df.loc[fill_mask, "Urbanization_national_ref"], errors="coerce") * alpha
			df.loc[fill_mask, "Urbanization_source"] = "National_urbanization_scaled_by_prov_ratio"
			df.loc[fill_mask, "Urbanization_is_national_proxy"] = 1
			df.loc[fill_mask, "Urbanization_is_imputed"] = 1
		if obs_proxy_mask.any():
			df.loc[obs_proxy_mask, "Urbanization_source"] = "Provincial_urbanization_share_observed"
			df.loc[obs_proxy_mask, "Urbanization_is_national_proxy"] = 0
			df.loc[obs_proxy_mask, "Urbanization_is_imputed"] = 0
	df = df.drop(columns=["Urbanization_prov"], errors="ignore")
	df = df.drop(columns=["Urbanization_national_ref"], errors="ignore")

	df["Urbanization"] = pd.to_numeric(df["Urbanization"], errors="coerce").clip(lower=0.01, upper=99.99)

	return df


def safe_log(s: pd.Series) -> pd.Series:
	vals = np.log(s.where(s > 0))
	return pd.Series(vals, index=s.index)


def add_kaya_features(panel: pd.DataFrame) -> pd.DataFrame:
	df = panel.copy()
	df["S"] = pd.to_numeric(df["Industry"], errors="coerce") / 100.0
	df["A"] = df["GDP"] / df["Population"]
	df["B"] = df["Energy"] / (df["GDP"] * df["S"])
	df["C"] = df["CO2"] / df["Energy"]

	df["ln_CO2"] = safe_log(df["CO2"])
	df["ln_Population"] = safe_log(df["Population"])
	df["ln_A"] = safe_log(df["A"])
	df["ln_S"] = safe_log(df["S"])
	df["ln_B"] = safe_log(df["B"])
	df["ln_C"] = safe_log(df["C"])
	return df


def log_mean(a: float, b: float) -> Optional[float]:
	if a is None or b is None:
		return None
	if a <= 0 or b <= 0:
		return None
	if abs(a - b) < 1e-12:
		return float(a)
	da = np.log(a)
	db = np.log(b)
	if abs(da - db) < 1e-12:
		return float(a)
	return float((a - b) / (da - db))


def compute_lmdi_time(panel: pd.DataFrame) -> pd.DataFrame:
	rows: List[dict] = []

	for province, g in panel.sort_values(["province", "year"]).groupby("province"):
		g = g.reset_index(drop=True)
		for i in range(1, len(g)):
			prev = g.iloc[i - 1]
			curr = g.iloc[i]

			e0, et = prev["CO2"], curr["CO2"]
			p0, pt = prev["Population"], curr["Population"]
			a0, at = prev["A"], curr["A"]
			s0, st = prev["S"], curr["S"]
			b0, bt = prev["B"], curr["B"]
			c0, ct = prev["C"], curr["C"]

			valid = all(
				pd.notna(v) and float(v) > 0
				for v in [e0, et, p0, pt, a0, at, s0, st, b0, bt, c0, ct]
			)
			if not valid:
				continue

			lm = log_mean(float(et), float(e0))
			if lm is None:
				continue

			d_p = lm * np.log(float(pt) / float(p0))
			d_a = lm * np.log(float(at) / float(a0))
			d_s = lm * np.log(float(st) / float(s0))
			d_b = lm * np.log(float(bt) / float(b0))
			d_c = lm * np.log(float(ct) / float(c0))
			d_co2 = float(et) - float(e0)
			resid = d_co2 - (d_p + d_a + d_s + d_b + d_c)
			resid_ratio_abs = abs(resid) / max(abs(d_co2), 1e-12)

			rows.append(
				{
					"province": province,
					"year": int(curr["year"]),
					"base_year": int(prev["year"]),
					"delta_CO2": d_co2,
					"delta_P": d_p,
					"delta_A": d_a,
					"delta_S": d_s,
					"delta_B": d_b,
					"delta_C": d_c,
					"lmdi_residual": resid,
					"resid_ratio": resid_ratio_abs,
					"lmdi_residual_abs_ratio": resid_ratio_abs,
				}
			)

	return pd.DataFrame(rows)


def attach_controls(panel: pd.DataFrame, controls: List[VariableSeries]) -> pd.DataFrame:
	df = panel.copy()
	for s in controls:
		if s.data.empty:
			continue
		df = df.merge(s.data, on="year", how="left")
		df[f"{s.name}_source"] = s.source
		df[f"{s.name}_is_national_proxy"] = 1

	if "Urbanization" not in df.columns:
		df["Urbanization"] = np.nan

	if "UrbanPopulation" in df.columns and "Population" in df.columns:
		calc = (df["UrbanPopulation"] / df["Population"]) * 100.0
		df["Urbanization"] = df["Urbanization"].fillna(calc)
		if "Urbanization_source" not in df.columns:
			df["Urbanization_source"] = "Derived_from_National_UrbanPopulation_over_Population"
			df["Urbanization_is_national_proxy"] = 1

	if "UrbanPopulation" in df.columns:
		df = df.drop(columns=["UrbanPopulation"], errors="ignore")
		df = df.drop(columns=["UrbanPopulation_source", "UrbanPopulation_is_national_proxy"], errors="ignore")

	return df


def fill_energy_with_interpolation_and_fit(
	panel: pd.DataFrame,
	national_energy_series: Optional[VariableSeries],
) -> pd.DataFrame:
	df = ensure_cols(
		panel.copy(),
		{
			"Energy": np.nan,
			"Energy_source": np.nan,
			"Energy_is_national_proxy": 0,
			"Energy_is_imputed": 0,
			"Energy_scale_factor": np.nan,
		},
	)

	if national_energy_series is None or national_energy_series.data.empty:
		return df

	national = national_energy_series.data.rename(columns={"Energy": "Energy_national"})
	df = df.merge(national, on="year", how="left")

	for year, idx in df.groupby("year").groups.items():
		year_int = int(cast(Any, year))
		year_idx = list(idx)
		year_mask = df.index.isin(year_idx)
		nat_vals = df.loc[year_idx, "Energy_national"].dropna()
		if nat_vals.empty:
			continue
		national_total = float(nat_vals.iloc[0])
		observed_mask = year_mask & df["Energy"].notna()
		missing_mask = year_mask & df["Energy"].isna()

		if year_int < 1997 or not observed_mask.any():
			weights = pd.to_numeric(df.loc[year_mask, "GDP"], errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)
			weight_sum = float(weights.sum())
			if weight_sum > 0:
				alloc_values = (weights.astype(float) / weight_sum).to_numpy(dtype=float) * national_total
				df.loc[year_mask, "Energy"] = alloc_values
				df.loc[year_mask, "Energy_source"] = "National_energy_allocated_by_GDP_share_pre1997"
			else:
				count_missing = int(year_mask.sum())
				if count_missing > 0:
					df.loc[year_mask, "Energy"] = national_total / count_missing
					df.loc[year_mask, "Energy_source"] = "National_energy_equal_split_fallback_pre1997"
			df.loc[year_mask, "Energy_is_national_proxy"] = 1
			df.loc[year_mask, "Energy_is_imputed"] = 1
			df.loc[year_mask, "Energy_scale_factor"] = 1.0
			continue

		if missing_mask.any():
			weights = pd.to_numeric(df.loc[missing_mask, "GDP"], errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)
			weight_sum = float(weights.sum())
			observed_sum = float(df.loc[observed_mask, "Energy"].sum(skipna=True))
			residual = national_total - observed_sum
			if weight_sum > 0:
				alloc = max(residual, 0.0) * (weights / weight_sum)
				df.loc[missing_mask, "Energy"] = alloc.values
				df.loc[missing_mask, "Energy_source"] = "National_energy_allocated_by_GDP_share_residual"
			else:
				count_missing = int(missing_mask.sum())
				if count_missing > 0:
					df.loc[missing_mask, "Energy"] = max(residual, 0.0) / count_missing
					df.loc[missing_mask, "Energy_source"] = "National_energy_equal_split_fallback"
			df.loc[missing_mask, "Energy_is_national_proxy"] = 1
			df.loc[missing_mask, "Energy_is_imputed"] = 1

		year_total = float(pd.to_numeric(df.loc[year_mask, "Energy"], errors="coerce").sum(skipna=True))
		if np.isfinite(year_total) and year_total > 0 and np.isfinite(national_total) and national_total > 0:
			scale = national_total / year_total
			if not np.isfinite(scale) or scale <= 0:
				scale = 1.0
			scaled_energy = pd.to_numeric(df.loc[year_mask, "Energy"], errors="coerce").to_numpy(dtype=float) * scale
			df.loc[year_mask, "Energy"] = scaled_energy
			df.loc[year_mask, "Energy_scale_factor"] = scale

	df["Energy"] = pd.to_numeric(df["Energy"], errors="coerce")
	df["Energy_is_national_proxy"] = pd.to_numeric(df["Energy_is_national_proxy"], errors="coerce").fillna(0)
	df["Energy_is_imputed"] = pd.to_numeric(df["Energy_is_imputed"], errors="coerce").fillna(0)
	return df


def allocate_missing_energy_from_national_proxy(panel: pd.DataFrame, source_name: str) -> pd.DataFrame:
	df = panel.copy()
	df = ensure_cols(
		df,
		{
			"Energy": np.nan,
			"Energy_source": np.nan,
			"Energy_is_national_proxy": np.nan,
			"Energy_is_imputed": 0,
		},
	)

	for year, idx in df.groupby("year").groups.items():
		year_idx = list(idx)
		nat_vals = df.loc[year_idx, "Energy_proxy"].dropna()
		if nat_vals.empty:
			continue
		nat_energy = float(nat_vals.iloc[0])
		year_mask = df.index.isin(year_idx)
		missing_mask = year_mask & df["Energy"].isna()
		if not missing_mask.any():
			continue

		year_gdp = pd.to_numeric(df.loc[year_mask, "GDP"], errors="coerce")
		year_industry = pd.to_numeric(df.loc[year_mask, "Industry"], errors="coerce") if "Industry" in df.columns else pd.Series(np.nan, index=df.index[year_mask])
		weights = year_gdp.loc[missing_mask].copy()
		if not year_industry.empty:
			industry_component = year_industry.loc[missing_mask].fillna(0.0)
			weights = weights * (1.0 + industry_component / 100.0)
		weights = weights.replace([np.inf, -np.inf], np.nan).fillna(0.0)
		weight_sum = float(weights.sum())
		if weight_sum > 0:
			weights = weights / weight_sum
		else:
			weights = pd.Series(0.0, index=df.loc[missing_mask].index)

		observed_sum = float(df.loc[year_mask & df["Energy"].notna(), "Energy"].sum(skipna=True))
		remaining = nat_energy - observed_sum
		allocatable = remaining if np.isfinite(remaining) and remaining > 0 else 0.0
		if weight_sum > 0:
			fill_values = allocatable * weights
			fill_source = source_name + "_residual_allocated_by_GDP_share"
		else:
			count_missing = int(missing_mask.sum())
			fill_values = pd.Series(allocatable / count_missing if count_missing > 0 else 0.0, index=df.loc[missing_mask].index)
			fill_source = source_name + "_equal_split_fallback"
		df.loc[missing_mask, "Energy"] = fill_values.values if hasattr(fill_values, "values") else fill_values
		df.loc[missing_mask, "Energy_source"] = fill_source
		df.loc[missing_mask, "Energy_is_national_proxy"] = 1
		df.loc[missing_mask, "Energy_is_imputed"] = 1

	df["Energy"] = pd.to_numeric(df["Energy"], errors="coerce")
	df["Energy_is_national_proxy"] = pd.to_numeric(df["Energy_is_national_proxy"], errors="coerce").fillna(0)
	df["Energy_is_imputed"] = pd.to_numeric(df["Energy_is_imputed"], errors="coerce").fillna(0)
	return df


def validate_unit_scale(panel: pd.DataFrame) -> Dict[str, float]:
	df = panel.copy()
	for col in ["CO2", "GDP", "Population", "Energy"]:
		if col not in df.columns:
			raise ValueError(f"Missing required column for unit validation: {col}")
		vals = pd.to_numeric(df[col], errors="coerce").dropna()
		if vals.empty:
			raise ValueError(f"No numeric values available for unit validation: {col}")

	means = {
		col: float(pd.to_numeric(df[col], errors="coerce").mean())
		for col in ["CO2", "GDP", "Energy"]
	}
	energy_over_gdp = means["Energy"] / means["GDP"] if np.isfinite(means["GDP"]) and means["GDP"] > 0 else np.nan
	co2_over_energy = means["CO2"] / means["Energy"] if np.isfinite(means["Energy"]) and means["Energy"] > 0 else np.nan
	if not np.isfinite(energy_over_gdp) or not (1e-8 < energy_over_gdp < 1e3):
		raise ValueError(f"Suspicious unit scale: Energy/GDP mean ratio = {energy_over_gdp}")
	if not np.isfinite(co2_over_energy) or not (1e-8 < co2_over_energy < 1e3):
		raise ValueError(f"Suspicious unit scale: CO2/Energy mean ratio = {co2_over_energy}")
	return {
		"CO2_mean": means["CO2"],
		"GDP_mean": means["GDP"],
		"Energy_mean": means["Energy"],
		"Energy_over_GDP_mean_ratio": float(energy_over_gdp),
		"CO2_over_Energy_mean_ratio": float(co2_over_energy),
	}


def collect_unit_consistency_warnings(panel: pd.DataFrame) -> List[str]:
	warnings: List[str] = []
	for col in ["CO2", "GDP", "Population", "Energy"]:
		if col not in panel.columns:
			warnings.append(f"Missing required column: {col}")
			continue
		vals = pd.to_numeric(panel[col], errors="coerce")
		if vals.notna().sum() == 0:
			warnings.append(f"Column has no numeric values: {col}")
		elif (vals <= 0).all():
			warnings.append(f"Column has no positive values: {col}")

	for ratio_col in ["B", "C"]:
		if ratio_col not in panel.columns:
			continue
		vals = pd.to_numeric(panel[ratio_col], errors="coerce")
		vals = vals[np.isfinite(vals) & (vals > 0)]
		if vals.empty:
			warnings.append(f"Ratio {ratio_col} has no positive finite values")
			continue
		med = float(vals.median())
		if med < 1e-8 or med > 1e8:
			warnings.append(f"Ratio {ratio_col} median out of expected range: {med:.3e}")

	unit_checks = [
		("Energy", "GDP", 1e-8, 1e2),
		("CO2", "Energy", 1e-8, 1e2),
		("CO2", "GDP", 1e-10, 1e2),
	]
	for numerator, denominator, lower, upper in unit_checks:
		if numerator not in panel.columns or denominator not in panel.columns:
			continue
		num_mean = float(pd.to_numeric(panel[numerator], errors="coerce").mean())
		den_mean = float(pd.to_numeric(panel[denominator], errors="coerce").mean())
		if not np.isfinite(num_mean) or not np.isfinite(den_mean) or den_mean <= 0:
			warnings.append(f"Invalid unit scale for {numerator}/{denominator}")
			continue
		ratio = num_mean / den_mean
		if ratio < lower or ratio > upper:
			warnings.append(
				f"Suspicious unit scale for {numerator}/{denominator}: {ratio:.3e} outside [{lower:.1e}, {upper:.1e}]"
			)

	return warnings


def main() -> None:
	ensure_output_dir()
	items = load_dataset_index()

	co2_panel = build_co2_panel(items)
	controls = build_national_controls(items)
	prov_gdp = read_provincial_wide_series(
		items,
		["省份维度", "经济相关", "省级GDP"],
		variable_name="GDP",
		source_name="Provincial_macro_GDP_1949_2024_observed",
	)
	prov_population = read_provincial_wide_series(
		items,
		["省份维度", "人口相关", "省级人口"],
		variable_name="Population",
		source_name="Provincial_macro_Population_1949_2024_observed",
	)
	energy_proxy_series = next((s for s in controls if s.name == "Energy"), None)
	industry_proxy_series = next((s for s in controls if s.name == "Industry"), None)
	urbanization_proxy_series = next((s for s in controls if s.name == "Urbanization"), None)
	non_energy_controls = [s for s in controls if s.name not in {"Energy", "Industry", "UrbanPopulation"}]

	panel = attach_controls(co2_panel, non_energy_controls)
	panel = merge_provincial_variable(panel, prov_gdp)
	panel = merge_provincial_variable(panel, prov_population)

	prov_industry = read_provincial_industry_share(items)
	if not prov_industry.empty:
		panel = panel.merge(prov_industry, on=["province", "year"], how="left")
	else:
		if "Industry" not in panel.columns:
			panel["Industry"] = np.nan
		if "Industry_source" not in panel.columns:
			panel["Industry_source"] = np.nan
		if "Industry_is_national_proxy" not in panel.columns:
			panel["Industry_is_national_proxy"] = np.nan
	if "Industry_is_imputed" not in panel.columns:
		panel["Industry_is_imputed"] = 0

	panel = fill_industry_with_interpolation_fit_and_anchor(panel, industry_proxy_series)

	prov_urbanization = read_provincial_urbanization_share(items)
	if not prov_urbanization.empty:
		panel = panel.merge(prov_urbanization, on=["province", "year"], how="left")
	if "Urbanization_is_imputed" not in panel.columns:
		panel["Urbanization_is_imputed"] = 0

	panel = fill_urbanization_with_interpolation_fit_and_anchor(panel, urbanization_proxy_series)

	prov_energy = read_provincial_energy_inventory(items)
	if not prov_energy.empty:
		panel = panel.merge(prov_energy, on=["province", "year"], how="left")
	else:
		if "Energy" not in panel.columns:
			panel["Energy"] = np.nan
		if "Energy_source" not in panel.columns:
			panel["Energy_source"] = np.nan
		if "Energy_is_national_proxy" not in panel.columns:
			panel["Energy_is_national_proxy"] = np.nan
	if "Energy_is_imputed" not in panel.columns:
		panel["Energy_is_imputed"] = 0

	panel = fill_energy_with_interpolation_and_fit(panel, energy_proxy_series)

	panel["Energy_is_national_proxy"] = panel["Energy_is_national_proxy"].fillna(0)
	panel["Energy_is_imputed"] = panel["Energy_is_imputed"].fillna(0)
	panel = add_kaya_features(panel)

	core_cols = [
		"province",
		"year",
		"CO2",
		"GDP",
		"Population",
		"Energy",
		"Industry",
		"S",
		"Urbanization",
		"A",
		"B",
		"C",
		"CO2_source",
	]
	panel_core = cast(pd.DataFrame, panel.loc[:, [c for c in core_cols if c in panel.columns]].copy())

	# Keep rows where CO2 exists; remaining variables are filled by available controls.
	panel_core = cast(
		pd.DataFrame,
		panel_core.sort_values(by=["year", "province"], kind="mergesort").reset_index(drop=True),
	)

	unit_diagnostics = validate_unit_scale(panel_core)
	lmdi_df = compute_lmdi_time(panel_core)
	unit_warnings = collect_unit_consistency_warnings(panel_core)
	lmdi_residual_ratio_stats = {}
	if not lmdi_df.empty and "lmdi_residual_abs_ratio" in lmdi_df.columns:
		lmdi_residual_ratio_stats = {
			"mean": float(lmdi_df["lmdi_residual_abs_ratio"].mean()),
			"median": float(lmdi_df["lmdi_residual_abs_ratio"].median()),
			"p90": float(lmdi_df["lmdi_residual_abs_ratio"].quantile(0.9)),
			"max": float(lmdi_df["lmdi_residual_abs_ratio"].max()),
		}

	panel_path = OUT_DIR / "panel_master.csv"
	lmdi_path = OUT_DIR / "lmdi_decomposition.csv"
	audit_path = OUT_DIR / "panel_source_audit.csv"
	summary_path = OUT_DIR / "panel_build_summary.json"

	panel_core.to_csv(panel_path, index=False, encoding="utf-8-sig")
	lmdi_df.to_csv(lmdi_path, index=False, encoding="utf-8-sig")

	audit_cols = [
		c
		for c in panel.columns
		if c.endswith("_source") or c.endswith("_is_national_proxy") or c.endswith("_is_imputed")
	]
	panel[["province", "year", "CO2"] + audit_cols].to_csv(audit_path, index=False, encoding="utf-8-sig")

	summary = {
		"panel_rows": int(len(panel_core)),
		"province_count": int(panel_core["province"].nunique()),
		"year_min": int(panel_core["year"].min()) if len(panel_core) else None,
		"year_max": int(panel_core["year"].max()) if len(panel_core) else None,
		"co2_source_counts": panel_core["CO2_source"].value_counts(dropna=False).to_dict(),
		"gdp_source_counts": panel["GDP_source"].value_counts(dropna=False).to_dict() if "GDP_source" in panel.columns else {},
		"population_source_counts": panel["Population_source"].value_counts(dropna=False).to_dict() if "Population_source" in panel.columns else {},
		"unit_diagnostics": unit_diagnostics,
		"unit_consistency_warnings": unit_warnings,
		"lmdi_residual_abs_ratio_stats": lmdi_residual_ratio_stats,
		"missing_ratio": {
			c: float(panel_core[c].isna().mean())
			for c in ["CO2", "GDP", "Population", "Energy", "Industry", "Urbanization", "A", "B", "C"]
			if c in panel_core.columns
		},
		"energy_source_counts": panel["Energy_source"].value_counts(dropna=False).to_dict()
		if "Energy_source" in panel.columns
		else {},
		"industry_source_counts": panel["Industry_source"].value_counts(dropna=False).to_dict()
		if "Industry_source" in panel.columns
		else {},
		"urbanization_source_counts": panel["Urbanization_source"].value_counts(dropna=False).to_dict()
		if "Urbanization_source" in panel.columns
		else {},
		"notes": [
			"CO2 uses MEIC 1990-2022 only; no NBS splice is used.",
			"GDP and Population use provincial observed series when available, with national macro series retained only as fallback.",
			"Industry is included as a structural effect in the Kaya identity; B is defined as energy intensity after structural adjustment so the five-factor identity remains exact.",
			"Energy uses provincial inventory (1997-2022) plus national-energy constraint: 1990-1996 are GDP-weighted allocations, and each year is rescaled to the national total.",
			"Industry uses provincial share series (observed mostly from 1996 onward), with logit interpolation/backcast and additive national anchor shift for imputed years.",
			"Urbanization uses provincial observed series when available and fills missing years by province-specific ratio to the national urbanization series.",
			"LMDI residual diagnostics are reported in resid_ratio, lmdi_residual_abs_ratio, and summary stats.",
		],
	}
	summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

	print("WROTE", panel_path)
	print("WROTE", lmdi_path)
	print("WROTE", audit_path)
	print("WROTE", summary_path)
	print("PANEL_ROWS", summary["panel_rows"], "PROVINCES", summary["province_count"], "YEARS", summary["year_min"], summary["year_max"])
	print("CO2_SOURCE_COUNTS", summary["co2_source_counts"])


if __name__ == "__main__":
	main()